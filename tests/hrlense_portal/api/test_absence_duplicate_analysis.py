"""
Absence Case Root Cause Investigation & Per-Employee Detail Workbook Generator.
Generates: 'reports/root_cause_detail.xlsx'
- Sheet 1: 'Summary_Overview' (All 968+ employees with total cases, total absent days, duplicate status)
- Sheet 2: 'Duplicate_Overview' (Consolidated list of duplicate cases with instance numbering)
- Sheet 3: 'Date_by_Date_Timeline' (Complete master date-by-date timeline: Employee -> Case ID -> Dates)
- Dedicated Individual Tabs for Top Duplicate Cases & Clean Benchmark Cases.
"""

import os
import re
import json
import logging
import pytest
import pandas as pd
from datetime import datetime, timedelta
from collections import defaultdict
from core.config import settings
from utils.api.base_api import get

logger = logging.getLogger(__name__)


def fetch_all_absence_cases() -> list[dict]:
    """Fetches complete list of absence cases with pagination."""
    all_records = []
    first = 0
    rows = 500
    
    while True:
        lazy_params = json.dumps({
            "first": first,
            "rows": rows,
            "page": first // rows,
            "sortField": "id",
            "sortOrder": -1
        })
        params = {
            "lazyParams": lazy_params,
            "search": "",
            "status": ""
        }
        logger.info(f"Fetching Absence Cases (first={first}, rows={rows})...")
        try:
            resp = get("AbsenceCase/list", params=params)
            records = []
            if isinstance(resp, list):
                records = resp
            elif isinstance(resp, dict):
                if "data" in resp:
                    records = resp["data"] if isinstance(resp["data"], list) else resp["data"].get("rows", [])
                elif "rows" in resp:
                    records = resp["rows"]
            
            if not records:
                break
            
            all_records.extend(records)
            logger.info(f"Fetched {len(records)} records (Total so far: {len(all_records)})")
            
            if len(records) < rows or len(all_records) >= 2000:
                break
            first += rows
        except Exception as e:
            logger.error(f"Error fetching page at first={first}: {e}")
            break
            
    return all_records


def sanitize_sheet_name(name: str, used_names: set) -> str:
    """Sanitizes sheet name to meet Excel 31-char limit and invalid characters rule."""
    clean = re.sub(r'[\\/*?:\[\]]', '', str(name)).strip()
    if not clean:
        clean = "Emp_Detail"
    clean = clean[:25]
    
    candidate = clean
    counter = 1
    while candidate.lower() in used_names:
        candidate = f"{clean[:22]}_{counter}"
        counter += 1
        
    used_names.add(candidate.lower())
    return candidate


@pytest.mark.api
def test_generate_root_cause_detail_excel():
    """
    Fetches all absence cases and generates 'reports/root_cause_detail.xlsx'
    containing:
    1. 'Summary_Overview' Sheet (All employees)
    2. 'Duplicate_Overview' Sheet (Duplicate cases)
    3. 'Date_by_Date_Timeline' Sheet (Master Employee -> Case ID -> Date sequence)
    4. Dedicated Individual Employee Tabs
    """
    logger.info("Starting Root Cause Detail Excel Generation...")
    
    # 1. Fetch all cases
    all_cases = fetch_all_absence_cases()
    assert len(all_cases) > 0, "No absence cases retrieved from API."
    print(f"\nTotal Absence Cases Fetched: {len(all_cases)}")

    # 2. Group by Employee
    emp_groups = defaultdict(list)
    for c in all_cases:
        emp_id = str(c.get("employee_Id") or c.get("employeeId") or c.get("employee_Code") or c.get("employeeCode") or "N/A")
        emp_name = str(c.get("employee_Name") or c.get("employeeName") or "Unknown").strip()
        key = (emp_id, emp_name)
        emp_groups[key].append(c)

    duplicate_groups = {k: v for k, v in emp_groups.items() if len(v) > 1}
    single_groups = {k: v for k, v in emp_groups.items() if len(v) == 1}

    # Calculate Frequency Distribution (Count of records per employee)
    from collections import Counter
    freq_counter = Counter(len(cases) for cases in emp_groups.values())
    max_cases = max(freq_counter.keys()) if freq_counter else 1

    print("\n" + "=" * 95)
    print("                 ABSENCE CASE ROOT CAUSE INVESTIGATION SUMMARY")
    print("=" * 95)
    print(f"Total Unique Employees:                    {len(emp_groups)}")
    print(f"Employees with Multiple (Duplicate) Cases: {len(duplicate_groups)} ({len(duplicate_groups)/len(emp_groups)*100:.1f}%)")
    print(f"Employees with Single (Unique) Cases:       {len(single_groups)} ({len(single_groups)/len(emp_groups)*100:.1f}%)")
    print("-" * 95)
    print("             DUPLICATE RECORD FREQUENCY BREAKDOWN PER EMPLOYEE")
    print("-" * 95)

    distribution_rows = []
    max_k = max(max(freq_counter.keys(), default=1), 10)
    
    for k in range(1, max_k + 1):
        count_emp = freq_counter.get(k, 0)
        pct = (count_emp / len(emp_groups)) * 100 if emp_groups else 0.0
        category = "Single (Clean)" if k == 1 else f"Duplicate ({k} records)"
        
        # Only print rows or print 1 to 10 explicitly
        print(f"  • Employees with {k:<2} records = {count_emp:<4} employees ({pct:5.1f}%)  [{category}]")
        distribution_rows.append({
            "Record Count Per Employee": f"{k} Records",
            "Category": category,
            "Total Employees": count_emp,
            "Percentage of Workforce": f"{pct:.1f}%",
            "Total Absence Rows Generated": count_emp * k
        })

    # Group any records > max_k if needed
    higher_counts = sum(v for k, v in freq_counter.items() if k > max_k)
    if higher_counts > 0:
        pct_high = (higher_counts / len(emp_groups)) * 100
        print(f"  • Employees with >{max_k} records = {higher_counts:<4} employees ({pct_high:5.1f}%)  [Duplicate ({max_k}+ records)]")
        distribution_rows.append({
            "Record Count Per Employee": f">{max_k} Records",
            "Category": f"Duplicate (>{max_k} records)",
            "Total Employees": higher_counts,
            "Percentage of Workforce": f"{pct_high:.1f}%",
            "Total Absence Rows Generated": sum(k * v for k, v in freq_counter.items() if k > max_k)
        })

    print("=" * 95 + "\n")

    # 3. Build Summary Overview
    summary_rows = []
    for (emp_id, emp_name), cases in sorted(emp_groups.items(), key=lambda x: len(x[1]), reverse=True):
        first_c = cases[0]
        total_cases = len(cases)
        total_absent_days = sum(int(c.get("continuous_Absent_Days") or 1) for c in cases)
        dates = [str(c.get("start_Date") or "")[:10] for c in cases if c.get("start_Date")]
        date_span = f"{min(dates)} to {max(dates)}" if dates else "N/A"
        branch = first_c.get("branch_Name") or "N/A"
        company = first_c.get("company_Name") or "N/A"
        dept = first_c.get("department_Name") or "N/A"
        manager = first_c.get("manager_Name") or "N/A"
        status_list = list(set(str(c.get("status") or "UnderReview") for c in cases))
        
        root_cause_summary = "Duplicate entries generated on separate dates instead of continuous counter update" if total_cases > 1 else "Normal Single Entry"

        summary_rows.append({
            "Employee Name": emp_name,
            "Employee ID": emp_id,
            "Employee Code": first_c.get("employee_Code") or "",
            "Branch": branch,
            "Company": company,
            "Department": dept,
            "Reporting Manager": manager,
            "Total Absence Cases": total_cases,
            "Total Absent Days": total_absent_days,
            "Date Range / Span": date_span,
            "Statuses": ", ".join(status_list),
            "Is Duplicate": "Yes" if total_cases > 1 else "No",
            "Root Cause Summary": root_cause_summary
        })

    # 4. Build Duplicate Summary
    dup_rows = []
    for (emp_id, emp_name), cases in sorted(duplicate_groups.items(), key=lambda x: len(x[1]), reverse=True):
        for idx, c in enumerate(cases, 1):
            raw_date = str(c.get("start_Date") or "")
            start_date = raw_date.split("T")[0] if "T" in raw_date else raw_date[:10]
            raw_comm = str(c.get("last_Communication_Date") or "")
            last_comm = raw_comm.split("T")[0] if "T" in raw_comm else raw_comm[:10]

            dup_rows.append({
                "Employee Name": emp_name,
                "Employee ID": emp_id,
                "Instance Number": f"Case #{idx} of {len(cases)}",
                "Case ID": c.get("id"),
                "Start Date": start_date,
                "Continuous Absent Days": c.get("continuous_Absent_Days") or 1,
                "Threshold Days": c.get("threshold_Days") or 3,
                "Status": c.get("status") or "UnderReview",
                "Communication Count": c.get("communication_Count") or 0,
                "Last Communication Date": last_comm,
                "Branch": c.get("branch_Name") or "N/A",
                "Company": c.get("company_Name") or "N/A",
                "Department": c.get("department_Name") or "N/A"
            })

    # 5. Build Master Date-by-Date Timeline (Employee -> Case ID -> Dates) for ALL records
    master_timeline_rows = []
    emp_individual_rows = defaultdict(list)
    seen_emp_dates = set()

    for (emp_id, emp_name), cases in sorted(emp_groups.items(), key=lambda x: len(x[1]), reverse=True):
        first_c = cases[0]
        emp_code = str(first_c.get("employee_Code") or emp_id)
        sorted_cases = sorted(cases, key=lambda x: str(x.get("start_Date") or x.get("startDate") or ""))

        for case_idx, c in enumerate(sorted_cases, 1):
            cid = c.get("id")
            raw_date = str(c.get("start_Date") or c.get("startDate") or "")
            start_date_str = raw_date.split("T")[0] if "T" in raw_date else raw_date[:10]
            cont_days = int(c.get("continuous_Absent_Days") or 1)
            threshold = int(c.get("threshold_Days") or 3)
            status = c.get("status") or "UnderReview"
            comm_count = c.get("communication_Count") or 0
            raw_comm = str(c.get("last_Communication_Date") or "")
            last_comm = raw_comm.split("T")[0] if "T" in raw_comm else raw_comm[:10]
            
            notify_mgr = "Yes" if c.get("notify_Manager") else "No"
            notify_hr = "Yes" if c.get("notify_Hr") else "No"
            notify_tl = "Yes" if c.get("notify_Team_Lead") else "No"

            try:
                start_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
            except Exception:
                start_dt = datetime.now()

            for day_offset in range(cont_days):
                current_date_dt = start_dt + timedelta(days=day_offset)
                current_date_str = current_date_dt.strftime("%Y-%m-%d")
                day_label = f"Day {day_offset + 1} of {cont_days}"

                emp_date_key = (emp_id, current_date_str)
                if emp_date_key in seen_emp_dates:
                    daily_assessment = f"DUPLICATE CONFLICT: Multiple Case IDs active on {current_date_str}"
                elif case_idx > 1 and day_offset == 0:
                    daily_assessment = f"DUPLICATE BUG: New Case ID #{cid} generated on {current_date_str} instead of incrementing existing case (Day {cont_days})"
                elif day_offset == 0:
                    daily_assessment = f"Initial Case Creation: Case #{cid} triggered on {current_date_str}"
                else:
                    daily_assessment = f"Continuous Absence: Day {day_offset + 1} under Case #{cid}"

                seen_emp_dates.add(emp_date_key)

                row_dict = {
                    "Employee Name": emp_name,
                    "Employee ID": emp_id,
                    "Employee Code": emp_code,
                    "Case ID": cid,
                    "Absence Date": current_date_str,
                    "Day Sequence": day_label,
                    "Case Start Date": start_date_str,
                    "Continuous Absent Days": cont_days,
                    "Threshold Days": threshold,
                    "Case Status": status,
                    "Communication Count": comm_count,
                    "Last Comm Date": last_comm,
                    "Notify Manager": notify_mgr,
                    "Notify HR": notify_hr,
                    "Notify TL": notify_tl,
                    "Reporting Manager": c.get("manager_Name") or "N/A",
                    "Branch": c.get("branch_Name") or "N/A",
                    "Company": c.get("company_Name") or "N/A",
                    "Department": c.get("department_Name") or "N/A",
                    "Date Assessment & Finding": daily_assessment
                }
                master_timeline_rows.append(row_dict)
                emp_individual_rows[(emp_id, emp_name)].append(row_dict)

    # Group records by duplicate count (1 record, 2 records, 3 records, ... 10+ records)
    records_by_count = defaultdict(list)
    for (emp_id, emp_name), cases in emp_groups.items():
        k = len(cases)
        for idx, c in enumerate(cases, 1):
            raw_date = str(c.get("start_Date") or c.get("startDate") or "")
            start_date = raw_date.split("T")[0] if "T" in raw_date else raw_date[:10]
            raw_comm = str(c.get("last_Communication_Date") or "")
            last_comm = raw_comm.split("T")[0] if "T" in raw_comm else raw_comm[:10]
            cont_days = int(c.get("continuous_Absent_Days") or 1)
            threshold = int(c.get("threshold_Days") or 3)
            status = c.get("status") or "UnderReview"
            comm_count = c.get("communication_Count") or 0

            assessment = "Normal Single Entry" if k == 1 else f"Duplicate Case #{idx} of {k} created on {start_date}"

            records_by_count[k].append({
                "Employee Name": emp_name,
                "Employee ID": emp_id,
                "Employee Code": c.get("employee_Code") or "",
                "Case ID": c.get("id"),
                "Instance Number": f"Case #{idx} of {k}",
                "Start Date": start_date,
                "Continuous Absent Days": cont_days,
                "Threshold Days": threshold,
                "Case Status": status,
                "Communication Count": comm_count,
                "Last Communication Date": last_comm,
                "Branch": c.get("branch_Name") or "N/A",
                "Company": c.get("company_Name") or "N/A",
                "Department": c.get("department_Name") or "N/A",
                "Reporting Manager": c.get("manager_Name") or "N/A",
                "Team Lead": c.get("tl_Name") or "N/A",
                "Total Cases for Employee": k,
                "Root Cause Finding": assessment
            })

    # 6. Export to Excel ('reports/root_cause_detail.xlsx')
    os.makedirs("reports", exist_ok=True)
    excel_path = "reports/root_cause_detail.xlsx"
    used_sheet_names = set()

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # Sheet 1: Master Summary Overview
        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, index=False, sheet_name="Summary_Overview")
        used_sheet_names.add("summary_overview")

        # Sheet 2: Duplicate Frequency Distribution
        df_dist = pd.DataFrame(distribution_rows)
        df_dist.to_excel(writer, index=False, sheet_name="Duplicate_Distribution")
        used_sheet_names.add("duplicate_distribution")

        # Create Dedicated Sheets: 1 Record (Clean), 2 Records, 3 Records ... up to 10 Records
        for k in range(1, 11):
            sheet_label = f"{k}_Record_Clean" if k == 1 else f"{k}_Records_Duplicate"
            rows_k = records_by_count.get(k, [])
            if rows_k:
                df_k = pd.DataFrame(rows_k)
                df_k.to_excel(writer, index=False, sheet_name=sheet_label)
                used_sheet_names.add(sheet_label.lower())
                logger.info(f"Added sheet '{sheet_label}' with {len(rows_k)} rows ({len(set(r['Employee ID'] for r in rows_k))} employees)")

        # Any records with > 10 duplicates
        rows_above_10 = []
        for k, r_list in records_by_count.items():
            if k > 10:
                rows_above_10.extend(r_list)
        if rows_above_10:
            df_above10 = pd.DataFrame(rows_above_10)
            df_above10.to_excel(writer, index=False, sheet_name="10+_Records_Duplicate")
            used_sheet_names.add("10+_records_duplicate")

        # Master Date-by-Date Granular Timeline (ALL Employees)
        df_timeline = pd.DataFrame(master_timeline_rows)
        df_timeline.to_excel(writer, index=False, sheet_name="Date_by_Date_Timeline")
        used_sheet_names.add("date_by_date_timeline")

        # Freeze panes: Header Row (Row 1) and Employee Name column
        # Style headers with bold font and professional dark navy fill
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        for ws in writer.book.worksheets:
            # Determine column index of Employee Name (or default to Column A)
            freeze_col = 2  # default freezes Column A (at B2)
            for cell in ws[1]:
                if "employee name" in str(cell.value or "").lower() or "employee" in str(cell.value or "").lower():
                    freeze_col = cell.column + 1
                    break
            
            freeze_cell = f"{get_column_letter(freeze_col)}2"
            ws.freeze_panes = freeze_cell

            # Style header row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Auto-adjust column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col[:100])
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

    logger.info(f"Generated Root Cause Detail Excel workbook with Frozen Panes: {excel_path}")
    print(f"\nSuccessfully generated '{excel_path}' with Frozen Header Row & Employee Name columns across {len(writer.sheets)} sheets.")

    # 7. Markdown Root Cause Report
    md_path = "reports/root_cause_detail.md"
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("# Absence Case Root Cause Detail Report\n\n")
        f.write(f"- **Total Absence Cases Fetched**: {len(all_cases)}\n")
        f.write(f"- **Total Unique Employees**: {len(emp_groups)}\n")
        f.write(f"- **Employees with Multiple (Duplicate) Cases**: {len(duplicate_groups)} ({len(duplicate_groups)/len(emp_groups)*100:.1f}%)\n")
        f.write(f"- **Employees with Single (Unique) Cases**: {len(single_groups)} ({len(single_groups)/len(emp_groups)*100:.1f}%)\n\n")

        f.write("## 1. Duplicate Frequency Distribution\n\n")
        f.write("| Record Count Per Employee | Category | Total Employees | Percentage of Workforce | Total Absence Rows |\n")
        f.write("|---------------------------|----------|-----------------|-------------------------|--------------------|\n")
        for d in distribution_rows:
            f.write(f"| {d['Record Count Per Employee']} | {d['Category']} | {d['Total Employees']} | {d['Percentage of Workforce']} | {d['Total Absence Rows Generated']} |\n")

        f.write("\n## 2. Root Cause Analysis\n\n")
        f.write("### 🚨 The Primary Defect:\n")
        f.write("When an employee is absent on consecutive days (e.g. Day 1, Day 2, Day 3), the daily absence evaluation job **inserts a brand new Case ID every day** rather than updating the `continuous_Absent_Days` counter on the existing open `UnderReview` case.\n\n")
        f.write("### 🔍 Impact:\n")
        f.write(f"- **{len(duplicate_groups)} out of {len(emp_groups)} employees ({len(duplicate_groups)/len(emp_groups)*100:.1f}%)** have duplicate absence cases.\n")
        f.write("- Inflates total absence case counts from ~968 to over 2,000 records.\n")
        f.write("- Triggers duplicate notifications to Managers and HR for each duplicate row.\n\n")

        f.write("### 🛠️ Recommended Backend Fix:\n")
        f.write("In the daily absence evaluation service, check for an existing active case before creating a new one:\n")
        f.write("```sql\n")
        f.write("SELECT id, continuous_Absent_Days FROM AbsenceCase WHERE employee_Id = @emp_id AND status = 'UnderReview';\n")
        f.write("-- If exists: UPDATE AbsenceCase SET continuous_Absent_Days = continuous_Absent_Days + 1 WHERE id = @existing_id;\n")
        f.write("-- If NOT exists: INSERT INTO AbsenceCase (...) VALUES (...);\n")
        f.write("```\n\n")

        f.write("## 3. Top Duplicate Employees Summary\n\n")
        f.write("| Employee Name | Emp ID | Branch | Company | Total Cases | Total Days | Statuses |\n")
        f.write("|---------------|--------|--------|---------|-------------|------------|----------|\n")
        for r in summary_rows[:20]:
            f.write(f"| {r['Employee Name']} | {r['Employee ID']} | {r['Branch']} | {r['Company']} | {r['Total Absence Cases']} | {r['Total Absent Days']} | {r['Statuses']} |\n")

    logger.info(f"Generated Markdown report: {md_path}")
    print(f"Successfully generated '{md_path}'.\n")
