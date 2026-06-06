import os
import re
import sys
import glob
from datetime import datetime
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.api.payroll_api import (
    get_branches,
    get_payroll_status,
    get_payroll_list,
    get_employee_detail,
    get_bank_detail,
    get_balance_leave
)

YEAR = 2026
MONTH = 4

def _missing_blockers(e: dict) -> list[str]:
    r  = e["record"]   # payroll list record  — fields 1-4
    d  = e["detail"]   # employee detail      — fields 5, 6, 7, 9
    b  = e["bank"]     # bank detail          — field 10
    lv = e["leave"]    # balance leave        — field 8

    missing = []

    # 1. Payroll Company
    if not r.get("payrollCompanyId"):
        missing.append("Payroll Company (payrollCompanyId is null/0)")

    # 2. Company
    if not r.get("companyId"):
        missing.append("Company (companyId is null/0)")

    # 3. Branch
    if not r.get("branchId"):
        missing.append("Branch (branchId is null/0)")

    # 4. Department
    if not r.get("departmentId"):
        missing.append("Department (departmentId is null/0)")

    # 5. Shift
    if not d.get("shift"):
        missing.append("Shift (shift is null/0)")

    # 6. Business Process
    if not (d.get("business_Process") or "").strip():
        missing.append("Business Process (business_Process is null/empty)")

    # 7. Date of Joining
    if not d.get("date_Of_Joining"):
        missing.append("Date of Joining (date_Of_Joining is null)")

    # 8. Balance Leave
    if not lv.get("balanceLeaves", []) and not lv.get("data", []):
        missing.append("Balance Leave (no balance leave records configured)")

    # 9. Salary
    if not (d.get("basic_salary") or 0) > 0:
        missing.append(
            f"Salary (basic_salary={d.get('basic_salary')}, "
            f"gross_Salary_Per_Month={d.get('gross_Salary_Per_Month')})"
        )

    # 10. Bank Details
    account = b.get("account_Number")
    ifsc    = b.get("ifsc_Code")
    if not account or not ifsc:
        parts = []
        if not account: parts.append("account_Number")
        if not ifsc:    parts.append("ifsc_Code")
        missing.append(f"Bank Details ({', '.join(parts)} is null/missing)")

    return missing

def scan_and_generate_text_report(target_branches=None) -> tuple[str, list]:
    """Scans all branches and generates a consolidated blocker text report."""
    print("==================================================")
    if target_branches:
        print(f"SCANNING TARGET BRANCHES: {', '.join(target_branches)}")
    else:
        print("SCANNING ALL SYSTEM BRANCHES FOR ALL EMPLOYEES")
    print("==================================================")
    
    branches = get_branches()
    if target_branches:
        filtered_branches = []
        for b in branches:
            b_name = b.get("branch_Name", "").lower()
            matched = False
            for tb in target_branches:
                # Handle common spelling variations for Bhubaneswar
                tb_clean = "bhubaneswar" if ("bhumb" in tb or "bhub" in tb or "bhumbhneshwar" in tb) else tb
                if tb_clean in b_name or tb in b_name:
                    matched = True
                    break
            if matched:
                filtered_branches.append(b)
        branches = filtered_branches
    
    pending_branches_count = 0
    total_records = 0
    all_reports = []
    branches_data = []
    
    # Track branches with payrolls
    for idx, b in enumerate(branches, 1):
        branch_id = b["id"]
        branch_name = b.get("branch_Name", f"Branch {branch_id}")
        company_name = b.get("company_Name", "Unknown Company")
        
        # Fetch ALL records (no status filter)
        try:
            response = get_payroll_list(year=YEAR, month=MONTH, branch_id=branch_id, status=None)
            records = response.get("data", [])
        except Exception as e:
            print(f"  Error fetching payroll list for branch {branch_name}: {e}")
            continue
            
        if len(records) > 0:
            pending_branches_count += 1
            total_records += len(records)
            print(f"\n[{pending_branches_count}] Branch '{branch_name}' ({company_name}) has {len(records)} EMPLOYEES!")
            
            # Gather details
            pending_data = []
            for emp in records:
                emp_id = emp["employeeId"]
                print(f"  - Gathering details for employee '{emp.get('employeeName')}' (ID: {emp_id})...")
                try:
                    detail = get_employee_detail(emp_id)
                except Exception:
                    detail = {}
                try:
                    bank = get_bank_detail(emp_id)
                except Exception:
                    bank = {}
                try:
                    leave = get_balance_leave(emp_id)
                except Exception:
                    leave = {}
                    
                pending_data.append({
                    "employeeId":   emp_id,
                    "employeeName": emp.get("employeeName", "Unknown").strip(),
                    "employeeCode": emp.get("employeeCode", ""),
                    "record":       emp,
                    "detail":       detail,
                    "bank":         bank,
                    "leave":        leave,
                })
                
            # Run blocker checks
            blocked_by = {}
            all_blocked = []
            
            for e in pending_data:
                missing = _missing_blockers(e)
                status = e["record"].get("status", "pending")
                name = f"{e['employeeName']} (code: {e['employeeCode']}) | Status: {status}"
                if missing:
                    all_blocked.append(name)
                    for m in missing:
                        field = m.split(" (")[0]
                        blocked_by.setdefault(field, []).append(name)
                        
            # Build report section
            lines = []
            lines.append("=" * 70)
            lines.append(f"ALL EMPLOYEE PAYROLL BLOCKER REPORT — {company_name} | {branch_name} | Month: {MONTH}/{YEAR} — {len(pending_data)} employees")
            lines.append("=" * 70)
            
            if all_blocked:
                lines.append(f"\n{len(all_blocked)}/{len(pending_data)} BLOCKED:")
                for i, name in enumerate(all_blocked, 1):
                    lines.append(f"  {i}. {name}")
                    
            lines.append("\nBLOCKER BREAKDOWN:")
            for field, names in sorted(blocked_by.items(), key=lambda x: -len(x[1])):
                lines.append(f"\n  {len(names)} missing '{field}':")
                for i, name in enumerate(names, 1):
                    lines.append(f"    {i}. {name}")
                    
            lines.append("=" * 70)
            report_section = "\n".join(lines)
            all_reports.append(report_section)
            
            # Store full branch/company data including all employees
            branches_data.append({
                'company_name': company_name,
                'branch_name': branch_name,
                'month': MONTH,
                'year': YEAR,
                'total_employees': len(pending_data),
                'employees': [
                    {
                        'name': e['employeeName'],
                        'code': e['employeeCode'],
                        'status': e['record'].get('status', 'pending'),
                        'missing': _missing_blockers(e)
                    } for e in pending_data
                ]
            })
            
    # Compile and write consolidated report
    if all_reports:
        consolidated_report = "\n\n".join(all_reports)
        os.makedirs("reports/payroll data report", exist_ok=True)
        report_path = f"reports/payroll data report/payroll_release_{datetime.today().strftime('%Y-%m-%d')}_all_branches.txt"
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(consolidated_report)
            
        print("\n==================================================")
        print(f"CONSOLIDATED REPORT WRITTEN: {report_path}")
        print("==================================================")
        return report_path, branches_data
    else:
        print("\n==================================================")
        print("NO RECORD DETECTED FOR ANY SYSTEM BRANCH!")
        print("==================================================")
        return "", []

def parse_consolidated_report(txt_file_path: str) -> list:
    """Parse the consolidated payroll blocker text report and extract data for all branches."""
    if not os.path.exists(txt_file_path):
        print(f"Error: Text report file not found at {txt_file_path}")
        return []
        
    with open(txt_file_path, 'r', encoding='utf-8') as f:
        content = f.read()
        
    # Split content by the separator lines (70 '=' signs)
    sections = re.split(r"={70}\s*\n(?=(?:PENDING|ALL EMPLOYEE) PAYROLL BLOCKER REPORT)", content)
    
    parsed_branches = []
    
    for sec in sections:
        sec = sec.strip()
        if not sec:
            continue
            
        # Parse header info
        header_pattern = r"(?:PENDING|ALL EMPLOYEE) PAYROLL BLOCKER REPORT — (.+?) \| (.+?) \| Month: (\d+)/(\d+) — (\d+) employees\s*\n={70}"
        header_match = re.search(header_pattern, sec)
        if not header_match:
            continue
            
        company_name = header_match.group(1).strip()
        branch_name = header_match.group(2).strip()
        month = int(header_match.group(3))
        year = int(header_match.group(4))
        total_employees = int(header_match.group(5))
        
        # Extract blocked employees
        blocked_section = re.search(r"(\d+)/(\d+) BLOCKED:\n(.*?)(?=BLOCKER BREAKDOWN:|$)", sec, re.DOTALL)
        
        blocked_employees = []
        if blocked_section:
            employees_text = blocked_section.group(3)
            emp_pattern = r"\d+\.\s+(.+?)\s+\(code:\s*(\w*)\)(?:\s*\|\s*Status:\s*(\w+))?"
            for match in re.finditer(emp_pattern, employees_text):
                name = match.group(1).strip()
                code = match.group(2).strip()
                status = match.group(3).strip() if match.group(3) else "pending"
                blocked_employees.append({
                    'Employee Name': name,
                    'Employee Code': code,
                    'Payroll Generation Status': status
                })
                
        # Extract blocker breakdown
        blocker_map = {}  # blocker_type -> set of employee codes
        breakdown_section = re.search(r"BLOCKER BREAKDOWN:(.+)$", sec, re.DOTALL)
        
        if breakdown_section:
            breakdown_text = breakdown_section.group(1)
            # Find each blocker type section
            blocker_pattern = r"\s*\d+\s+missing\s+'([^']+)':\n(.*?)(?=\n\s*\d+\s+missing|\n={70}|\Z)"
            for match in re.finditer(blocker_pattern, breakdown_text, re.DOTALL):
                blocker_type = match.group(1).strip()
                names_text = match.group(2)
                
                codes = set()
                emp_pattern = r"\d+\.\s+(.+?)\s+\(code:\s*(\w*)\)"
                for emp_match in re.finditer(emp_pattern, names_text):
                    codes.add(emp_match.group(2).strip())
                    
                blocker_map[blocker_type] = codes
                
        parsed_branches.append({
            'company_name': company_name,
            'branch_name': branch_name,
            'month': month,
            'year': year,
            'total_employees': total_employees,
            'blocked_employees': blocked_employees,
            'blocker_map': blocker_map
        })
        
    return parsed_branches

def generate_branch_excel(branch_name: str, companies_in_branch: list, output_dir: str):
    """Generate or update a single-branch dedicated Excel validation report with one sheet per company."""
    os.makedirs(output_dir, exist_ok=True)
    
    def sanitize_filename(name):
        return re.sub(r'[\\/*?:"<>|]', '', name).strip()
        
    branch_clean = sanitize_filename(branch_name)
    excel_filename = f"{branch_clean} blocker report.xlsx"
    excel_path = os.path.join(output_dir, excel_filename)
    
    # Load existing workbook if it exists, otherwise create a new one
    if os.path.exists(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path)
            print(f"  - Loaded existing branch blocker Excel: {excel_path}")
        except Exception as e:
            print(f"  - Failed to load existing workbook {excel_path} ({e}). Creating new.")
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    
    for branch in companies_in_branch:
        company_clean = sanitize_filename(branch['company_name'])
        # Sheet titles can be max 31 characters
        sheet_title = company_clean[:31]
        
        # If the sheet already exists in the workbook, remove it to overwrite it with fresh data
        if sheet_title in wb.sheetnames:
            wb.remove(wb[sheet_title])
            
        ws = wb.create_sheet(title=sheet_title)
        
        # Styles
        red_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
        red_font = Font(color='C00000', bold=True)
        green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        green_font = Font(color='375623', bold=False)
        
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Blockers (standard field columns in order)
        field_columns = [
            'Payroll Company',
            'Company',
            'Branch',
            'Department',
            'Shift',
            'Business Process',
            'Date of Joining',
            'Balance Leave',
            'Salary',
            'Bank Details'
        ]
        
        # Map blocker names in text report to column names
        blocker_to_field = {
            'Payroll Company': 'Payroll Company',
            'Company': 'Company',
            'Branch': 'Branch',
            'Department': 'Department',
            'Shift': 'Shift',
            'Business Process': 'Business Process',
            'Date of Joining': 'Date of Joining',
            'Balance Leave': 'Balance Leave',
            'Salary': 'Salary',
            'Bank Details': 'Bank Details'
        }
        
        # Write Title block
        TITLE = f"ALL EMPLOYEE PAYROLL BLOCKER REPORT  —  {branch['company_name']}  —  {branch['branch_name']}  —  Month: {branch['month']}/{branch['year']}"
        ws.append([TITLE])
        ws.merge_cells("A1:M1")
        ws["A1"].font = Font(bold=True, size=11, color="1F4E79")
        ws.row_dimensions[1].height = 24
        
        ws.append([]) # Blank spacer
        ws.row_dimensions[2].height = 8
        
        # Headers
        headers = ['S.No', 'Employee Name', 'Employee Code', 'Payroll Generation Status'] + field_columns
        ws.append(headers)
        ws.row_dimensions[3].height = 22
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            
        # Write employee rows
        emps_to_write = branch.get('employees')
        if emps_to_write is None:
            # Recreate from blocked_employees and blocker_map (old parser fallback)
            emps_to_write = []
            for emp in branch.get('blocked_employees', []):
                emp_code = emp['Employee Code']
                missing = []
                for blocker_type, codes in branch.get('blocker_map', {}).items():
                    if emp_code in codes:
                        missing.append(blocker_type)
                emps_to_write.append({
                    'name': emp['Employee Name'],
                    'code': emp_code,
                    'status': emp.get('Payroll Generation Status', 'pending'),
                    'missing': missing
                })
                
        for row_num, emp in enumerate(emps_to_write, 4):
            emp_code = emp['code']
            emp_status = emp.get('status', 'pending')
            
            row_data = [
                row_num - 3,            # S.No
                emp['name'],            # Name
                emp_code,               # Code
                emp_status.title()      # Payroll Generation Status
            ]
            
            # Map blocker statuses (YES/NO)
            for field in field_columns:
                found_blocker = False
                for m in emp['missing']:
                    if field.lower() in m.lower():
                        found_blocker = True
                        break
                row_data.append('NO' if found_blocker else 'YES')
                
            ws.append(row_data)
            ws.row_dimensions[row_num].height = 18
            
            # Style the cells in this row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border = thin_border
                cell.font = Font(size=9)
                
                if col_idx <= 4:
                    cell.alignment = Alignment(horizontal='center' if col_idx != 2 else 'left', vertical='center')
                    if col_idx == 4:
                        if cell.value == 'Pending':
                            cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                            cell.font = Font(size=9, color='B27A00', bold=True)
                        else:
                            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                            cell.font = Font(size=9, color='375623')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # Apply color mapping
                    if cell.value == 'NO':
                        cell.fill = red_fill
                        cell.font = red_font
                    else:
                        cell.fill = green_fill
                        cell.font = green_font
                        
        # Autofit column widths
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col[0].column == 1:
                ws.column_dimensions[col_letter].width = 6
            elif col[0].column == 2:
                ws.column_dimensions[col_letter].width = 25
            elif col[0].column == 3:
                ws.column_dimensions[col_letter].width = 15
            elif col[0].column == 4:
                ws.column_dimensions[col_letter].width = 22
            else:
                ws.column_dimensions[col_letter].width = 18
            
    # Save the branch workbook once all company sheets are added
    try:
        wb.save(excel_path)
        print(f"  - Successfully generated/updated branch blocker Excel: {excel_path} (Companies: {len(companies_in_branch)})")
    except PermissionError:
        from datetime import datetime
        ts = datetime.now().strftime('%H%M%S')
        excel_path_ts = excel_path.replace('.xlsx', f'_{ts}.xlsx')
        wb.save(excel_path_ts)
        print(f"  - Saved branch blocker Excel (with timestamp lock-avoidance): {excel_path_ts}")

def convert_text_to_excel_reports(txt_file_path: str):
    """Parses a specific text report and generates/updates Excel files in 'reports/payroll blocker report'."""
    if not txt_file_path or not os.path.exists(txt_file_path):
        print(f"Error: Blocker text report file not found at '{txt_file_path}'")
        return
        
    print(f"Parsing blocker text report: {txt_file_path}")
    branches_data = parse_consolidated_report(txt_file_path)
    print(f"Parsed {len(branches_data)} branches with blocker data.")
    
    if not branches_data:
        print("No blocker data found in text report.")
        return
        
    # Output directory
    output_dir = "reports/payroll blocker report"
    os.makedirs(output_dir, exist_ok=True)
    print(f"Generating/Updating branch-wise Excel files inside '{output_dir}':")
    
    # Group by branch (normalize casing to prevent overwriting files on Windows)
    grouped_by_branch = {}
    for b in branches_data:
        b_name = b['branch_name'].title()
        if b_name not in grouped_by_branch:
            grouped_by_branch[b_name] = []
        grouped_by_branch[b_name].append(b)
        
    for branch_name, companies in grouped_by_branch.items():
        generate_branch_excel(branch_name, companies, output_dir)
        
    print("All branch-wise blocker Excel reports successfully updated.")

def find_latest_text_report() -> str:
    """Locate the latest consolidated report in the default folders."""
    txt_sources = [
        "reports/payroll data report",
        "reports"
    ]
    for src in txt_sources:
        if not os.path.exists(src):
            continue
        files = glob.glob(os.path.join(src, "*all_branches.txt"))
        if not files:
            files = glob.glob(os.path.join(src, "payroll_release_*.txt"))
        if files:
            files.sort(key=os.path.getmtime, reverse=True)
            return files[0]
    return ""

def main():
    # Support filtering by branches, e.g. --branches="Bhubaneswar,Noida,Greater Noida"
    target_branches = None
    for arg in sys.argv:
        if arg.startswith("--branches="):
            target_branches = [b.strip().lower() for b in arg.split("=")[1].split(",")]

    # If passed '--convert-existing', run the Excel converter using the latest existing text report
    if len(sys.argv) > 1 and sys.argv[1] == "--convert-existing":
        latest_txt = find_latest_text_report()
        if not latest_txt:
            print("Error: No existing blocker report text files found.")
            sys.exit(1)
        convert_text_to_excel_reports(latest_txt)
        sys.exit(0)
        
    # Default sequential execution:
    # 1. Run the blocker scanner to generate the text report and gather all branch data
    report_path, branches_data = scan_and_generate_text_report(target_branches)
    
    # 2. Generate Excel reports using the full data (including all employees)
    if branches_data:
        output_dir = "reports/payroll blocker report"
        os.makedirs(output_dir, exist_ok=True)
        print(f"Generating/Updating branch-wise Excel files inside '{output_dir}':")
        
        # Group by branch
        grouped_by_branch = {}
        for b in branches_data:
            b_name = b['branch_name'].title()
            if b_name not in grouped_by_branch:
                grouped_by_branch[b_name] = []
            grouped_by_branch[b_name].append(b)
            
        for branch_name, companies in grouped_by_branch.items():
            generate_branch_excel(branch_name, companies, output_dir)
            
        print("All branch-wise blocker Excel reports successfully updated.")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("Execution failed:", e)
