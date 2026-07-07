import openpyxl
import os
import glob

xlsx_files = glob.glob("reports/payroll_validation_report_*.xlsx")
if not xlsx_files:
    print("No validation report xlsx found.")
    exit(1)

# Sort to get the latest one
latest_xlsx = max(xlsx_files, key=os.path.getmtime)
print(f"Loading {latest_xlsx}...")

wb = openpyxl.load_workbook(latest_xlsx)
print("Sheets in workbook:", wb.sheetnames)

for sheetname in ["T0 Field Detail", "T5 Difference"]:
    sheet = wb[sheetname]
    print(f"\n--- Sheet: {sheetname} ---")
    print(f"Dimensions: {sheet.dimensions}")

    # Print row 3 (which has headers)
    print("Row 3 (Headers):")
    row3_vals = [sheet.cell(row=3, column=c).value for c in range(1, min(sheet.max_column + 1, 45))]
    print(row3_vals)

    # Print row 4 (sub-headers or first employee)
    print("Row 4:")
    row4_vals = [sheet.cell(row=4, column=c).value for c in range(1, min(sheet.max_column + 1, 45))]
    print(row4_vals)

    # Print row 5
    print("Row 5:")
    row5_vals = [sheet.cell(row=5, column=c).value for c in range(1, min(sheet.max_column + 1, 45))]
    print(row5_vals)
