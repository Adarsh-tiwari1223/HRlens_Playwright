import os
import re
import glob
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_consolidated_report(txt_file_path):
    """Parse the consolidated payroll blocker text report and extract data for all branches."""
    if not os.path.exists(txt_file_path):
        print(f"Error: Text report file not found at {txt_file_path}")
        return []
        
    with open(txt_file_path, 'r', encoding='utf-8') as f:
        content = f.read()
        
    # Split content by the separator lines (70 '=' signs)
    sections = re.split(r"={70}\s*\n(?=PENDING PAYROLL BLOCKER REPORT)", content)
    
    parsed_branches = []
    
    for sec in sections:
        sec = sec.strip()
        if not sec:
            continue
            
        # Parse header info
        header_pattern = r"PENDING PAYROLL BLOCKER REPORT — (.+?) \| (.+?) \| Month: (\d+)/(\d+) — (\d+) employees\s*\n={70}"
        header_match = re.search(header_pattern, sec)
        if not header_match:
            continue
            
        company_name = header_match.group(1).strip()
        branch_name = header_match.group(2).strip()
        month = int(header_match.group(3))
        year = int(header_match.group(4))
        total_employees = int(header_match.group(5))
        
        # Extract blocked employees
        blocked_section = re.search(r"(\d+)/(\d+) BLOCKED:\n(.*?)(?=BLOCKER BREAKDOWN:|$)", sec, re.DOTALL)
        
        blocked_employees = []
        if blocked_section:
            employees_text = blocked_section.group(3)
            emp_pattern = r"\d+\.\s+(.+?)\s+\(code:\s*(\w*)\)"
            for match in re.finditer(emp_pattern, employees_text):
                name = match.group(1).strip()
                code = match.group(2).strip()
                blocked_employees.append({
                    'Employee Name': name,
                    'Employee Code': code
                })
                
        # Extract blocker breakdown
        blocker_map = {}  # blocker_type -> set of employee codes
        breakdown_section = re.search(r"BLOCKER BREAKDOWN:(.+)$", sec, re.DOTALL)
        
        if breakdown_section:
            breakdown_text = breakdown_section.group(1)
            # Find each blocker type section
            blocker_pattern = r"\s*\d+\s+missing\s+'([^']+)':\n(.*?)(?=\n\s*\d+\s+missing|\n={70}|\Z)"
            for match in re.finditer(blocker_pattern, breakdown_text, re.DOTALL):
                blocker_type = match.group(1).strip()
                names_text = match.group(2)
                
                codes = set()
                emp_pattern = r"\d+\.\s+(.+?)\s+\(code:\s*(\w*)\)"
                for emp_match in re.finditer(emp_pattern, names_text):
                    codes.add(emp_match.group(2).strip())
                    
                blocker_map[blocker_type] = codes
                
        parsed_branches.append({
            'company_name': company_name,
            'branch_name': branch_name,
            'month': month,
            'year': year,
            'total_employees': total_employees,
            'blocked_employees': blocked_employees,
            'blocker_map': blocker_map
        })
        
    return parsed_branches

def generate_individual_excel(branch, output_dir):
    """Generate a single-branch dedicated Excel validation report in the output directory."""
    os.makedirs(output_dir, exist_ok=True)
    
    # Safe slug for filenames
    company_slug = re.sub(r'[^a-zA-Z0-9]', '_', branch['company_name'])
    branch_slug = re.sub(r'[^a-zA-Z0-9]', '_', branch['branch_name'])
    
    # Strip consecutive underscores
    company_slug = re.sub(r'_+', '_', company_slug).strip('_')
    branch_slug = re.sub(r'_+', '_', branch_slug).strip('_')
    
    excel_filename = f"{company_slug}_{branch_slug}_payroll_blocker_report.xlsx"
    excel_path = os.path.join(output_dir, excel_filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Report"
    
    # Styles
    red_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
    red_font = Font(color='C00000', bold=True)
    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    green_font = Font(color='375623', bold=False)
    
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Blockers (standard field columns in order)
    field_columns = [
        'Payroll Company',
        'Company',
        'Branch',
        'Department',
        'Shift',
        'Business Process',
        'Date of Joining',
        'Balance Leave',
        'Salary',
        'Bank Details'
    ]
    
    # Map blocker names in text report to column names
    blocker_to_field = {
        'Payroll Company': 'Payroll Company',
        'Company': 'Company',
        'Branch': 'Branch',
        'Department': 'Department',
        'Shift': 'Shift',
        'Business Process': 'Business Process',
        'Date of Joining': 'Date of Joining',
        'Balance Leave': 'Balance Leave',
        'Salary': 'Salary',
        'Bank Details': 'Bank Details'
    }
    
    # Write Title block
    TITLE = f"PENDING PAYROLL BLOCKER REPORT  —  {branch['company_name']}  —  {branch['branch_name']}  —  Month: {branch['month']}/{branch['year']}"
    ws.append([TITLE])
    ws.merge_cells("A1:M1")
    ws["A1"].font = Font(bold=True, size=11, color="1F4E79")
    ws.row_dimensions[1].height = 24
    
    ws.append([]) # Blank spacer
    ws.row_dimensions[2].height = 8
    
    # Headers
    headers = ['S.No', 'Employee Name', 'Employee Code'] + field_columns
    ws.append(headers)
    ws.row_dimensions[3].height = 22
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        
    # Write employee rows
    for row_num, emp in enumerate(branch['blocked_employees'], 4):
        emp_code = emp['Employee Code']
        
        row_data = [
            row_num - 3,            # S.No
            emp['Employee Name'],    # Name
            emp_code                 # Code
        ]
        
        # Map blocker statuses (YES/NO)
        for field in field_columns:
            found_blocker = False
            for blocker_type, codes in branch['blocker_map'].items():
                mapped_field = blocker_to_field.get(blocker_type, blocker_type)
                if mapped_field == field and emp_code in codes:
                    found_blocker = True
                    break
            row_data.append('NO' if found_blocker else 'YES')
            
        ws.append(row_data)
        ws.row_dimensions[row_num].height = 18
        
        # Style the cells in this row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            cell.font = Font(size=9)
            
            if col_idx <= 3:
                cell.alignment = Alignment(horizontal='center' if col_idx != 2 else 'left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # Apply color mapping
                if cell.value == 'NO':
                    cell.fill = red_fill
                    cell.font = red_font
                else:
                    cell.fill = green_fill
                    cell.font = green_font
                    
    # Autofit column widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col[0].column == 1:
            ws.column_dimensions[col_letter].width = 6
        elif col[0].column == 2:
            ws.column_dimensions[col_letter].width = 25
        elif col[0].column == 3:
            ws.column_dimensions[col_letter].width = 15
        else:
            ws.column_dimensions[col_letter].width = 18
            
    # Save the branch workbook
    try:
        wb.save(excel_path)
        print(f"  - Generated separate blocker Excel: {excel_path}")
    except PermissionError:
        # Fallback with timestamp in case the standard name is locked/open
        from datetime import datetime
        ts = datetime.now().strftime('%H%M%S')
        excel_path_ts = excel_path.replace('.xlsx', f'_{ts}.xlsx')
        wb.save(excel_path_ts)
        print(f"  - Generated separate blocker Excel (with timestamp lock-avoidance): {excel_path_ts}")

def convert_latest_consolidated_report():
    """Locate the latest consolidated report, parse it, and generate separate Excel files in 'reports/payroll blocker report'."""
    # Lookup in 'reports/payroll data report' first, fallback to 'reports'
    txt_sources = [
        "reports/payroll data report",
        "reports"
    ]
    
    latest_txt = None
    for src in txt_sources:
        if not os.path.exists(src):
            continue
        files = glob.glob(os.path.join(src, "*all_branches.txt"))
        if not files:
            files = glob.glob(os.path.join(src, "payroll_release_*.txt"))
        if files:
            files.sort(key=os.path.getmtime, reverse=True)
            latest_txt = files[0]
            break
            
    if not latest_txt:
        print("Error: No pending blocker report text files found.")
        return
        
    print(f"Parsing latest report: {latest_txt}")
    branches_data = parse_consolidated_report(latest_txt)
    print(f"Parsed {len(branches_data)} branches with pending payrolls.")
    
    if not branches_data:
        print("No pending payroll data found in text report.")
        return
        
    # Output directory
    output_dir = "reports/payroll blocker report"
    os.makedirs(output_dir, exist_ok=True)
    print(f"Generating individual Excel files inside '{output_dir}':")
    
    for branch in branches_data:
        generate_individual_excel(branch, output_dir)
        
    print("All individual blocker Excel reports successfully generated.")

if __name__ == "__main__":
    convert_latest_consolidated_report()
