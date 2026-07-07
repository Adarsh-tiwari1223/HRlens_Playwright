import openpyxl
import os

def find_blockers():
    file_path = "reports/payroll blocker report/Agra blocker report.xlsx"
    if not os.path.exists(file_path):
        print(f"Error: {file_path} does not exist.")
        return
        
    wb = openpyxl.load_workbook(file_path)
    print(f"Loaded {file_path}. Sheets: {wb.sheetnames}\n")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"==================================================")
        print(f"Sheet: {sheet_name}")
        print(f"==================================================")
        
        # Row 1 is Title
        title_row = [cell.value for cell in ws[1] if cell.value is not None]
        print(f"Title: {title_row}")
        
        # Find headers in row 3
        headers = [cell.value for cell in ws[3]]
        # Clean trailing None in headers
        while headers and headers[-1] is None:
            headers.pop()
        print(f"Headers: {headers}")
        
        # Check rows starting from row 4
        data_rows = list(ws.iter_rows(values_only=True))[3:]
        blocked_employees = []
        
        for idx, row in enumerate(data_rows, start=4):
            # Check if this row is the "No blocked employees found" row
            if row[0] == "No blocked employees found. All employees passed payroll blocker checks.":
                print("No blocked employees found in this sheet.")
                continue
                
            # If the row is empty or doesn't have employee name
            if not row or len(row) < 3 or row[1] is None:
                continue
                
            # Check for 'NO' in any cell
            no_fields = []
            for col_idx, val in enumerate(row):
                if col_idx < len(headers):
                    header = headers[col_idx]
                else:
                    header = f"Col {col_idx+1}"
                
                if val == "NO":
                    no_fields.append(header)
            
            if no_fields:
                emp_name = row[1]
                emp_code = row[2]
                status = row[3]
                blocked_employees.append({
                    "row": idx,
                    "name": emp_name,
                    "code": emp_code,
                    "status": status,
                    "blocked_reasons": no_fields,
                    "row_values": [val for val in row[:len(headers)]]
                })
        
        print(f"Found {len(blocked_employees)} employees with 'NO' values:")
        for emp in blocked_employees:
            print(f"  Row {emp['row']}: Name: {emp['name']} (Code: {emp['code']}), Status: {emp['status']}")
            print(f"    Failed Checks ('NO'): {emp['blocked_reasons']}")
            print(f"    Full Data: {emp['row_values']}")
        print()

if __name__ == "__main__":
    find_blockers()
