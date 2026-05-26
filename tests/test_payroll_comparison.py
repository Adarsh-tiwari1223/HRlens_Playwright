import os
import re
import difflib
import pytest
import pandas as pd
from collections import defaultdict
from datetime import datetime
from utils.api.payroll_api import get_released_payroll, find_payroll_company_id

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
YEAR         = 2026
MONTH        = 4
REPORTS_DIR  = "reports"
FUZZY_CUTOFF = 0.85
PAYROLL_DATA = "testdata/static/PayrollData"

# (excel_filename, sheet_name, company_name_fragment, display_company, display_branch,
#  data_start_row, employee_name_col, employee_code_col, field_map)
PAYROLL_CONFIGS = [
    ("Originator-Salary-April'2026.xlsx",   "APR'26", "Originator", "Originator Informatics Pvt. Ltd.", "Varanasi (TEK)", 7, 4, 7, "wide"),
    ("Code Crewzs- Salary-April'2026.xlsx", "Apr-26", "Code Crewz", "Code Crewzs",                     "Varanasi",       7, 3, 6, "standard"),
    ("Infoserv- Salary-April'2026.xlsx",    "APR-26", "Infoserv",   "Infoserv",                        "Varanasi",       7, 3, 6, "standard"),
    ("INIT- Salary-April'2026.xlsx",        "APR'26", "INIT",       "INIT",                            "Varanasi",       6, 3, 6, "init"),
    ("Jobvritta- Salary-April'2026.xlsx",   "APR-26", "Jobvritta",  "Jobvritta",                       "Varanasi",       7, 4, 7, "wide"),
]

# (col_index, label, api_field, is_numeric, needs_confirmation)
FIELD_MAPS = {
    # Originator and Jobvritta include company/present-day columns before Salary Days.
    "wide": [
    (9,  "Salary Days",               "paidDays",             True, False),
    (15, "Basic Amt. (Earned)",        "basic",                True, False),
    (16, "H.R.A Amt. (Earned)",        "hra",                  True, False),
    (17, "Conv. Amt. (Earned)",        "conveyance_Allowance", True, False),
    (18, "INCENTIVE",                  "incentive",            True, False),
    (20, "DEDUCTIONS P.F.",            "employee_PF",          True, False),
    (21, "DEDUCTIONS E.S.I.",          "esic_Employee",        True, False),
    (23, "DEDUCTIONS T.D.S.",          "tds",                  True, False),
    (24, "TOT. Ded.",                  "totalDeductions",      True, False),
    (25, "Net Payable",                "netSalary",            True, False),
    (26, "Reserve Leave after Apr'26", "balance_Leave",        True, True),
    ],
    # Code Crewzs and Infoserv have name/code one column earlier.
    "standard": [
    (8,  "Salary Days",               "paidDays",             True, False),
    (14, "Basic Amt. (Earned)",        "basic",                True, False),
    (15, "H.R.A Amt. (Earned)",        "hra",                  True, False),
    (16, "Conv. Amt. (Earned)",        "conveyance_Allowance", True, False),
    (17, "INCENTIVE",                  "incentive",            True, False),
    (19, "DEDUCTIONS P.F.",            "employee_PF",          True, False),
    (20, "DEDUCTIONS E.S.I.",          "esic_Employee",        True, False),
    (22, "DEDUCTIONS T.D.S.",          "tds",                  True, False),
    (23, "TOT. Ded.",                  "totalDeductions",      True, False),
    (24, "Net Payable",                "netSalary",            True, False),
    (25, "Reserve Leave after Apr'26", "balance_Leave",        True, True),
    ],
    # INIT does not have the separate Present Days column, so earned fields start earlier.
    "init": [
    (7,  "Salary Days",               "paidDays",             True, False),
    (13, "Basic Amt. (Earned)",        "basic",                True, False),
    (14, "H.R.A Amt. (Earned)",        "hra",                  True, False),
    (15, "Conv. Amt. (Earned)",        "conveyance_Allowance", True, False),
    (16, "INCENTIVE",                  "incentive",            True, False),
    (18, "DEDUCTIONS P.F.",            "employee_PF",          True, False),
    (19, "DEDUCTIONS E.S.I.",          "esic_Employee",        True, False),
    (21, "DEDUCTIONS T.D.S.",          "tds",                  True, False),
    (22, "TOT. Ded.",                  "totalDeductions",      True, False),
    (23, "Net Payable",                "netSalary",            True, False),
    (24, "Reserve Leave after Apr'26", "balance_Leave",        True, True),
    ],
}

# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def normalize_code(val) -> str:
    s = str(val).strip()
    return s[:-2] if s.endswith(".0") else s


def normalize_name(val) -> str:
    return str(val).strip().upper()


def safe_float(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, float) and pd.isna(val):
        return 0.0
    try:
        return round(float(str(val).strip()), 2)
    except (ValueError, TypeError):
        return 0.0


def is_blank(val) -> bool:
    if val is None:
        return True
    if isinstance(val, float) and pd.isna(val):
        return True
    return str(val).strip() == ""


def read_excel(path: str, sheet: str, data_start_row: int) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet, header=None)
    return df.iloc[data_start_row:].reset_index(drop=True)


def build_api_lookup(api_records: list):
    composite_lookup, name_lookup = {}, {}
    for rec in api_records:
        name = normalize_name(rec.get("employeeName", ""))
        code = normalize_code(rec.get("employeeCode", ""))
        if name or code:
            composite_lookup[(name, code)] = rec
        if name:
            name_lookup[name] = rec
    return composite_lookup, name_lookup


def find_api_record(excel_name, excel_code, composite_lookup, name_lookup):
    rec = composite_lookup.get((excel_name, excel_code))
    if rec:
        return rec, "EXACT_COMPOSITE"
    rec = name_lookup.get(excel_name)
    if rec:
        return rec, "EXACT_NAME"
    if excel_name:
        matches = difflib.get_close_matches(excel_name, list(name_lookup.keys()), n=1, cutoff=FUZZY_CUTOFF)
        if matches:
            return name_lookup[matches[0]], f"FUZZY({matches[0]})"
    return None, None


def compare_fields(raw_row, excel_name, excel_code, api_rec, match_type, field_map) -> list:
    results = []
    for col_idx, label, api_field, is_numeric, needs_confirm in field_map:
        excel_val = raw_row.iloc[col_idx] if col_idx < len(raw_row) else None
        if needs_confirm:
            results.append({
                "Employee Name": excel_name, "Employee Code": excel_code,
                "Match Type": match_type,    "Field Name": label,
                "Excel Value": safe_float(excel_val), "API Value": api_rec.get(api_field, "—"),
                "Status": "Skipped", "Remarks": "Needs Business Confirmation",
            })
            continue
        api_val   = api_rec.get(api_field)
        excel_num = safe_float(excel_val)
        api_num   = safe_float(api_val)
        passed    = excel_num == api_num
        results.append({
            "Employee Name": excel_name, "Employee Code": excel_code,
            "Match Type": match_type,    "Field Name": label,
            "Excel Value": excel_num,    "API Value": api_num,
            "Status": "Pass" if passed else "Fail",
            "Remarks": f"API field: {api_field}",
        })
    return results


def run_comparison(df, composite_lookup, name_lookup, name_col: int, code_col: int,
                   field_map: list) -> list:
    detail_rows, matched_keys = [], set()

    for _, raw_row in df.iterrows():
        raw_name = raw_row.iloc[name_col]
        raw_code = raw_row.iloc[code_col]
        if is_blank(raw_name) and is_blank(raw_code):
            continue

        excel_name = normalize_name(raw_name) if not is_blank(raw_name) else ""
        excel_code = normalize_code(raw_code) if not is_blank(raw_code) else ""
        api_rec, match_type = find_api_record(excel_name, excel_code, composite_lookup, name_lookup)

        if api_rec is None:
            net_col = next((col for col, label, *_ in field_map if label == "Net Payable"), None)
            parent_salary = safe_float(raw_row.iloc[net_col]) if net_col is not None and net_col < len(raw_row) else "—"
            detail_rows.append({
                "Employee Name": str(raw_name).strip(), "Employee Code": excel_code,
                "Match Type": "—", "Field Name": "—",
                "Excel Value": "—", "API Value": "—",
                "Status": "Not Found in API",
                "Remarks": "No API record matched (exact or fuzzy)",
                "Parent Salary": parent_salary,
            })
            continue

        api_key = (normalize_name(api_rec.get("employeeName", "")),
                   normalize_code(api_rec.get("employeeCode", "")))
        matched_keys.add(api_key)
        detail_rows.extend(compare_fields(raw_row, str(raw_name).strip(), excel_code,
                                          api_rec, match_type, field_map))

    for (api_name, api_code), rec in composite_lookup.items():
        if (api_name, api_code) not in matched_keys:
            detail_rows.append({
                "Employee Name": rec.get("employeeName", ""), "Employee Code": api_code,
                "Match Type": "—", "Field Name": "—",
                "Excel Value": "—", "API Value": "—",
                "Status": "Not Found in Excel",
                "Remarks": "Employee exists in API but not in Excel",
            })

    return detail_rows


def build_wide_rows(detail_rows: list, field_map: list):
    emp_fields, emp_match_type, emp_order = defaultdict(dict), {}, []

    for r in detail_rows:
        if r["Status"] in ("Not Found in API", "Not Found in Excel"):
            continue
        key = (r["Employee Name"], r["Employee Code"])
        if key not in emp_fields:
            emp_order.append(key)
        emp_fields[key][r["Field Name"]] = r
        emp_match_type[key] = r["Match Type"]

    field_labels = [label for (_, label, _, _, needs_confirm) in field_map if not needs_confirm]

    wide_rows = []
    for key in emp_order:
        name, code = key
        row = {"Employee Name": name, "Employee Code": code, "Match Type": emp_match_type.get(key, "—")}
        any_fail = False
        for label in field_labels:
            fr = emp_fields[key].get(label)
            if fr:
                excel_v, api_v = fr["Excel Value"], fr["API Value"]
                try:
                    diff = round(float(excel_v) - float(api_v), 2)
                except (TypeError, ValueError):
                    diff = "—"
                if fr["Status"] == "Fail":
                    any_fail = True
            else:
                excel_v = api_v = diff = "—"
            row[f"{label} Excel"] = excel_v
            row[f"{label} API"]   = api_v
            row[f"{label} Diff"]  = diff
        row["Final Status"] = "Fail" if any_fail else "Pass"
        wide_rows.append(row)

    return wide_rows, field_labels


def build_tables(detail_rows: list, api_total: int = 0):
    exact_paid, exact_diff = [], []
    fuzzy_paid, fuzzy_diff = [], []
    spelling_rows, not_found, not_in_excel = [], [], []
    seen_spelling = set()
    emp_fields, emp_match_type = defaultdict(dict), {}

    for r in detail_rows:
        if r["Status"] == "Not Found in API":
            not_found.append({
                "Employee Name": r["Employee Name"],
                "Employee Code": r["Employee Code"],
                "Parent Salary": r.get("Parent Salary", "—"),
            })
            continue
        if r["Status"] == "Not Found in Excel":
            not_in_excel.append({"Employee Name": r["Employee Name"], "Employee Code": r["Employee Code"]})
            continue
        key = (r["Employee Name"], r["Employee Code"])
        emp_fields[key][r["Field Name"]] = r
        emp_match_type[key] = r["Match Type"]

    for key, fields in emp_fields.items():
        name, code  = key
        match_type  = emp_match_type[key]
        is_fuzzy    = "FUZZY" in match_type
        net_row     = fields.get("Net Payable")
        excel_net   = net_row["Excel Value"] if net_row else "—"
        api_net     = net_row["API Value"]   if net_row else "—"
        net_match   = net_row and net_row["Status"] == "Pass"
        entry       = {"Employee Name": name, "Employee Code": code, "Parent Salary": excel_net, "Payroll Salary": api_net}

        if is_fuzzy:
            if key not in seen_spelling:
                seen_spelling.add(key)
                m = re.search(r"FUZZY\((.+)\)", match_type)
                spelling_rows.append({"Parent Column (Excel)": name, "Payroll Column (API)": m.group(1) if m else match_type, "Type": "Spelling"})
            if net_match:
                fuzzy_paid.append(entry)
            else:
                try:
                    diff = round(float(excel_net) - float(api_net), 2)
                except (TypeError, ValueError):
                    diff = "—"
                fuzzy_diff.append({**entry, "Difference": diff})
        else:
            if net_match:
                exact_paid.append(entry)
            else:
                try:
                    diff = round(float(excel_net) - float(api_net), 2)
                except (TypeError, ValueError):
                    diff = "—"
                exact_diff.append({**entry, "Difference": diff})

    n_exact         = len(exact_paid) + len(exact_diff)
    n_fuzzy         = len(seen_spelling)
    n_total_matched = len(emp_fields)
    n_salary_match  = len(exact_paid) + len(fuzzy_paid)
    n_salary_diff   = len(exact_diff) + len(fuzzy_diff)
    n_not_api       = len(not_found)
    n_not_excel     = len(not_in_excel)

    t1 = [
        {"Category": "Excel Records (Source)",    "Count": n_total_matched + n_not_api},
        {"Category": "API Records (Payroll)",      "Count": api_total},
        {"Category": "---",                        "Count": "---"},
        {"Category": "Matched (Exact)",            "Count": n_exact},
        {"Category": "Total Matched",              "Count": n_total_matched},
        {"Category": "Matched (Fuzzy / Spelling)", "Count": n_fuzzy},
        {"Category": "---",                        "Count": "---"},
        {"Category": "Salary Match (= Paid)",      "Count": n_salary_match},
        {"Category": "Salary Diff  (!= Paid)",     "Count": n_salary_diff},
        {"Category": "---",                        "Count": "---"},
        {"Category": "Unmatched (Not in API)",     "Count": n_not_api},
        {"Category": "Unmatched (Not in Excel)",   "Count": n_not_excel},
        {"Category": "Total Unmatched",            "Count": n_not_api + n_not_excel},
    ]
    t2 = [
        {"Status": "Salary Match (Exact)", "Count": len(exact_paid)},
        {"Status": "Salary Diff  (Exact)", "Count": len(exact_diff)},
        {"Status": "---",                  "Count": "---"},
        {"Status": "Total Salary Match",   "Count": n_salary_match},
        {"Status": "Total Salary Diff",    "Count": n_salary_diff},
        {"Status": "Total Matched",        "Count": n_total_matched},
    ]

    all_diff = exact_diff + fuzzy_diff
    for i, r in enumerate(spelling_rows, 1): r["#"] = i
    for i, r in enumerate(exact_paid,    1): r["#"] = i
    for i, r in enumerate(all_diff,      1): r["#"] = i
    for i, r in enumerate(not_found,     1): r["#"] = i
    for i, r in enumerate(not_in_excel,  1): r["#"] = i

    return t1, t2, spelling_rows, exact_paid, all_diff, not_found, not_in_excel


def save_workbook(wide_rows, field_labels, t1, t2, t3, t4, t5, t6, t7,
                  company: str, branch: str, slug: str) -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(REPORTS_DIR, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = f"{REPORTS_DIR}/payroll_validation_{slug}_{ts}.xlsx"

    wb         = openpyxl.Workbook()
    wb.remove(wb.active)
    MONTH_NAME = datetime(YEAR, MONTH, 1).strftime("%B %Y").upper()
    TITLE      = f"PAYROLL MAPPING REPORT  —  {company}  —  {branch}  —  {MONTH_NAME}"

    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    SEP_FILL  = PatternFill("solid", fgColor="D9E1F2")
    FAIL_FILL = PatternFill("solid", fgColor="FFE0E0")
    PASS_FILL = PatternFill("solid", fgColor="E2EFDA")
    WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
    bdr       = Border(*[Side(style="thin")] * 0)  # rebuilt below
    thin      = Side(style="thin")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def make_sheet(title):
        ws = wb.create_sheet(title=title)
        ws.append([TITLE])
        ws["A1"].font = Font(bold=True, size=13)
        ws.append([])
        return ws

    def write_header(ws, cols):
        ws.append(cols)
        row = ws.max_row
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill, cell.font = HDR_FILL, HDR_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = bdr

    def write_row(ws, values, fill=None):
        ws.append(values)
        row = ws.max_row
        for c in range(1, len(values) + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = bdr
            cell.alignment = Alignment(horizontal="left")
            if fill:
                cell.fill = fill

    def write_total(ws, count):
        ws.append([])
        ws.append(["Total:", count])
        row = ws.max_row
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)

    def autofit(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 50)

    # T0
    ws0 = make_sheet("T0 Field Detail")
    ws0.append(["TABLE 0: FIELD-LEVEL DETAIL (Wide View)"])
    ws0[f"A{ws0.max_row}"].font = Font(bold=True, size=12)
    fixed = ["Employee Name", "Employee Code", "Match Type"]
    field_cols = [f"{l} {s}" for l in field_labels for s in ("Excel", "API", "Diff")]
    all_cols = fixed + field_cols + ["Final Status"]
    write_header(ws0, all_cols)
    fs_col = len(all_cols)
    for r in wide_rows:
        vals = [r["Employee Name"], r["Employee Code"], r["Match Type"]]
        for l in field_labels:
            vals += [r[f"{l} Excel"], r[f"{l} API"], r[f"{l} Diff"]]
        vals.append(r["Final Status"])
        write_row(ws0, vals)
    for ri in range(5, ws0.max_row + 1):
        cell = ws0.cell(row=ri, column=fs_col)
        cell.fill = PASS_FILL if cell.value == "Pass" else (FAIL_FILL if cell.value == "Fail" else cell.fill)
    autofit(ws0)

    # T1
    ws1 = make_sheet("T1 Summary")
    ws1.append(["TABLE 1: SUMMARY"])
    ws1[f"A{ws1.max_row}"].font = Font(bold=True, size=12)
    write_header(ws1, ["Category", "Count"])
    for r in t1:
        is_sep  = r["Category"] == "---"
        is_bold = r["Category"].startswith(("Total", "Excel", "API"))
        write_row(ws1, [r["Category"], r["Count"]], SEP_FILL if is_sep else None)
        if is_bold:
            row = ws1.max_row
            ws1.cell(row=row, column=1).font = Font(bold=True)
            ws1.cell(row=row, column=2).font = Font(bold=True)
    autofit(ws1)

    # T2
    ws2 = make_sheet("T2 Difference Table")
    ws2.append(["TABLE 2: DIFFERENCE TABLE"])
    ws2[f"A{ws2.max_row}"].font = Font(bold=True, size=12)
    write_header(ws2, ["Status", "Count"])
    for r in t2:
        is_sep  = r["Status"] == "---"
        is_bold = r["Status"].startswith("Total")
        write_row(ws2, [r["Status"], r["Count"]], SEP_FILL if is_sep else None)
        if is_bold:
            row = ws2.max_row
            ws2.cell(row=row, column=1).font = Font(bold=True)
            ws2.cell(row=row, column=2).font = Font(bold=True)
    autofit(ws2)

    # T3
    ws3 = make_sheet("T3 Spelling Mismatch")
    ws3.append(["TABLE 3: SPELLING MISMATCH (Fuzzy Matched)"])
    ws3[f"A{ws3.max_row}"].font = Font(bold=True, size=12)
    write_header(ws3, ["#", "Parent Column (Excel)", "Payroll Column (API)", "Type"])
    for r in t3:
        write_row(ws3, [r["#"], r["Parent Column (Excel)"], r["Payroll Column (API)"], r["Type"]])
    write_total(ws3, len(t3))
    autofit(ws3)

    # T4
    ws4 = make_sheet("T4 Matched Exact")
    ws4.append(["TABLE 4: MATCHED (Exact — Salary Match)"])
    ws4[f"A{ws4.max_row}"].font = Font(bold=True, size=12)
    write_header(ws4, ["#", "Employee Name", "Employee Code", "Parent Salary", "Payroll Salary"])
    for r in t4:
        write_row(ws4, [r["#"], r["Employee Name"], r["Employee Code"], r["Parent Salary"], r["Payroll Salary"]])
    write_total(ws4, len(t4))
    autofit(ws4)

    # T5
    ws5 = make_sheet("T5 Difference")
    ws5.append(["TABLE 5: DIFFERENCE (Parent Salary != Payroll Salary)"])
    ws5[f"A{ws5.max_row}"].font = Font(bold=True, size=12)
    write_header(ws5, ["#", "Employee Name", "Employee Code", "Parent Salary", "Payroll Salary", "Difference"])
    for r in t5:
        write_row(ws5, [r["#"], r["Employee Name"], r["Employee Code"], r["Parent Salary"], r["Payroll Salary"], r["Difference"]], FAIL_FILL)
    write_total(ws5, len(t5))
    autofit(ws5)

    # T6
    ws6 = make_sheet("T6 Not in Payroll")
    ws6.append(["TABLE 6: NOT IN PAYROLL (In Excel, No API Match)"])
    ws6[f"A{ws6.max_row}"].font = Font(bold=True, size=12)
    write_header(ws6, ["#", "Employee Name", "Employee Code", "Parent Salary"])
    for r in t6:
        write_row(ws6, [r["#"], r["Employee Name"], r["Employee Code"], r["Parent Salary"]])
    write_total(ws6, len(t6))
    autofit(ws6)

    # T7
    ws7 = make_sheet("T7 In API Not in Excel")
    ws7.append(["TABLE 7: IN API BUT NOT IN EXCEL"])
    ws7[f"A{ws7.max_row}"].font = Font(bold=True, size=12)
    write_header(ws7, ["#", "Employee Name (API)", "Employee Code (API)"])
    for r in t7:
        write_row(ws7, [r["#"], r["Employee Name"], r["Employee Code"]], WARN_FILL)
    write_total(ws7, len(t7))
    autofit(ws7)

    wb.save(path)
    print(f"\nReport saved: {path}")
    return path


# ---------------------------------------------------------------------------
# PARAMETRIZED TEST
# ---------------------------------------------------------------------------

@pytest.mark.api
@pytest.mark.regression
@pytest.mark.parametrize(
    "excel_filename,sheet_name,company_fragment,display_company,display_branch,"
    "data_start_row,name_col,code_col,field_map_key",
    PAYROLL_CONFIGS,
    ids=[cfg[3] for cfg in PAYROLL_CONFIGS],
)
def test_payroll_excel_vs_api(excel_filename, sheet_name, company_fragment,
                               display_company, display_branch, data_start_row,
                               name_col, code_col, field_map_key):
    excel_file = os.path.join(PAYROLL_DATA, excel_filename)
    if not os.path.exists(excel_file):
        pytest.skip(f"Excel file not found: {excel_file}")
    field_map = FIELD_MAPS.get(field_map_key)
    assert field_map, f"Unknown payroll field map: {field_map_key}"

    try:
        payroll_company_id = find_payroll_company_id(company_fragment)
    except ValueError as exc:
        pytest.fail(f"Company lookup failed: {exc}")

    try:
        api_response = get_released_payroll(YEAR, MONTH, payroll_company_id)
    except Exception as exc:
        pytest.fail(f"API call failed: {exc}")

    api_records = api_response.get("data", [])
    assert api_records, f"API returned empty data for company '{display_company}'"

    df = read_excel(excel_file, sheet_name, data_start_row)
    composite_lookup, name_lookup = build_api_lookup(api_records)
    detail_rows = run_comparison(df, composite_lookup, name_lookup, name_col, code_col, field_map)

    wide_rows, field_labels = build_wide_rows(detail_rows, field_map)
    t1, t2, t3, t4, t5, t6, t7 = build_tables(detail_rows, api_total=len(api_records))

    slug = re.sub(r"[^a-zA-Z0-9]", "_", display_company)[:30]
    report_path = save_workbook(wide_rows, field_labels, t1, t2, t3, t4, t5, t6, t7,
                                display_company, display_branch, slug)

    print(f"\n{'='*60}")
    print(f"  {display_company}  |  {display_branch}  |  {YEAR}-{MONTH:02d}")
    print(f"{'='*60}")
    print(f"  Excel Records   : {t1[0]['Count']}")
    print(f"  API Records     : {t1[1]['Count']}")
    print(f"  Total Matched   : {t1[4]['Count']}")
    print(f"    Exact         : {t1[3]['Count']}")
    print(f"    Fuzzy         : {t1[5]['Count']}")
    print(f"  Salary Match    : {t1[7]['Count']}")
    print(f"  Salary Diff     : {t1[8]['Count']}")
    print(f"  Not in API      : {t1[10]['Count']}")
    print(f"  Not in Excel    : {t1[11]['Count']}")
    print(f"  Total Unmatched : {t1[12]['Count']}")
    print(f"{'='*60}")
    print(f"  Report: {report_path}")

    if t5:
        print(f"\nDifferences (first 30 of {len(t5)}):")
        for r in t5[:30]:
            print(f"  [{r['Employee Code']}] {r['Employee Name']:<30} "
                  f"| Excel={r['Parent Salary']}  API={r['Payroll Salary']}  Diff={r['Difference']}")

    assert not t5, (
        f"{display_company}: {len(t5)} employee(s) have Net Payable mismatch. "
        "See T5 Difference sheet for details."
    )
