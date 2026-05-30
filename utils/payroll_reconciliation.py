import os
import re
import json
import shutil
import difflib
import logging
import openpyxl
import pandas as pd
from datetime import datetime
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Target payroll fields mapped to API fields
TARGET_FIELDS = {
    "paidDays": ["salary days", "payable days", "paid days", "total days", "present days", "days paid", "actual working days", "working days", "days"],
    "basic": ["basic", "basic salary", "basic amt", "basic+da salary", "basic+da", "basic da", "stipend"],
    "hra": ["hra", "h.r.a.", "hra amt", "house rent", "house rent allowance", "h.r.a. amt."],
    "conveyance_Allowance": ["conv", "conv.", "conveyance", "conveyance allowance", "conv amt", "conv. amt."],
    "incentive": ["incentive", "spcl alw", "special", "special allowance", "special alw", "spcl.", "special alw.", "incentives", "allowance", "spcl alw."],
    "employee_PF": ["pf", "p.f.", "provident", "provident fund", "employee pf", "deductions pf", "deductions p.f."],
    "esic_Employee": ["esi", "e.s.i.", "esic", "employee esi", "deductions esi", "deductions e.s.i.", "esic employee"],
    "tds": ["tds", "t.d.s.", "tax", "deductions tds", "deductions t.d.s.", "tds amt"],
    "netSalary": ["net payable", "net salary", "payable", "net payable amt", "net payable salary", "payable amt", "net amount", "total net", "net  payable"],
    "balance_Leave": ["reserve leave", "leave balance", "earned leave after", "casual leave after", "reserve leave after", "leave", "earned leave", "el", "cl"]
}

class StructuralValidationError(Exception):
    """Raised when the Excel file structure is invalid or unrecognized."""
    pass

def clean_val(val) -> str:
    """Consistently treat empty/None/NA values and return a stripped string."""
    if val is None:
        return ""
    if isinstance(val, float) and pd.isna(val):
        return ""
    s = str(val).strip()
    if s.lower() in ("nan", "none", "na", "n/a", "—", "-", "null"):
        return ""
    return s

def normalize_name(val) -> str:
    """Trim and lowercase name for flexible matching."""
    s = clean_val(val)
    # Replace duplicate whitespaces
    return " ".join(s.lower().split())

def normalize_code(val) -> str:
    """Convert employee code to string and clean formatting."""
    s = clean_val(val)
    if s.endswith(".0"):
        s = s[:-2]
    return s

def safe_float(val) -> float:
    """Clean and parse a numeric field safely, handling currency symbols and commas."""
    s = clean_val(val)
    if not s:
        return 0.0
    # Remove commas, dollar, rupee, and other currency symbols
    s = re.sub(r"[^\d\.\-]", "", s)
    try:
        return round(float(s), 2)
    except (ValueError, TypeError):
        return 0.0

def normalize_header_text(text: str) -> str:
    """Normalize header text by lowercasing and removing punctuation for comparison."""
    if not text:
        return ""
    t = str(text).lower().strip()
    t = re.sub(r'[\s\n\r]+', ' ', t)
    # Strip dots, hyphens, and other punctuation, but preserve plus, letters, numbers, and spaces
    t = re.sub(r'[^\w\s\+\-]', '', t)
    return " ".join(t.split())

class PayrollReconciliationAgent:
    def __init__(self, file_path: str, sheet_name: str):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.df = None
        self.header_rows = []
        self.merged_headers = []
        self.company_fragment = ""
        self.display_company = ""
        self.display_branch = "Varanasi" # Default fallback
        self.employee_name_col = -1
        self.employee_code_col = -1
        self.hrlense_id_col = -1
        self.net_salary_col = -1
        self.col_mapping = {} # logical_field_name -> (col_index, header_text)
        self.status = "PASS"
        self.duplicates = []
        self.mismatches = []
        self.header_validation = {}

    def scan_file(self) -> dict:
        """Step 1 & 2: Read workbook, detect layout dynamically, and map columns."""
        logger.info(f"Scanning sheet '{self.sheet_name}' in file '{self.file_path}'")
        try:
            self.df = pd.read_excel(self.file_path, sheet_name=self.sheet_name, header=None)
        except Exception as e:
            raise StructuralValidationError(f"Could not read sheet '{self.sheet_name}': {e}")

        # 1. Detect actual header rows based on content score
        self._detect_headers()
        
        # 2. Extract company name and find API fragment
        self._detect_company()

        # 3. Detect employee name, code, and net salary columns
        self._detect_core_columns()

        # 4. Map other logical columns
        self._map_logical_columns()

        # 5. Validate core structures
        self._validate_structure()

        return {
            "header_rows": self.header_rows,
            "company_fragment": self.company_fragment,
            "display_company": self.display_company,
            "employee_name_col": self.employee_name_col,
            "employee_code_col": self.employee_code_col,
            "net_salary_col": self.net_salary_col,
            "mapping": {k: {"col_index": v[0], "header": v[1]} for k, v in self.col_mapping.items()}
        }

    def _detect_headers(self):
        """Find the header rows by scoring the first 15 rows of the worksheet."""
        row_scores = []
        max_rows_to_scan = min(15, len(self.df))
        
        key_terms = ["name", "code", "salary", "basic", "hra", "conv", "pf", "esi", "tds", "payable", "days", "stipend", "uan", "pan", "father"]
        
        for r_idx in range(max_rows_to_scan):
            row = self.df.iloc[r_idx]
            score = 0
            for val in row:
                if pd.isna(val) or val is None:
                    continue
                val_str = str(val).lower()
                for term in key_terms:
                    if term in val_str:
                        score += 1
                        break
            row_scores.append((r_idx, score))

        # Find row with max score
        best_row, max_score = max(row_scores, key=lambda x: x[1])
        if max_score < 2:
            raise StructuralValidationError(f"No valid payroll header rows detected in sheet '{self.sheet_name}'.")

        # Determine if consecutive rows around best_row should be merged
        header_candidates = [best_row]
        
        # Check row before
        if best_row > 0:
            before_score = next(s for r, s in row_scores if r == best_row - 1)
            if before_score >= 2:
                header_candidates.append(best_row - 1)
        
        # Check row after
        if best_row < max_rows_to_scan - 1:
            after_score = next(s for r, s in row_scores if r == best_row + 1)
            if after_score >= 2:
                header_candidates.append(best_row + 1)

        self.header_rows = sorted(list(set(header_candidates)))
        
        # Merge consecutive header cells
        num_cols = len(self.df.columns)
        self.merged_headers = []
        
        for c_idx in range(num_cols):
            parts = []
            for r_idx in self.header_rows:
                cell_val = self.df.iloc[r_idx, c_idx]
                cell_str = clean_val(cell_val).replace("\n", " ").strip()
                if cell_str and cell_str not in parts:
                    parts.append(cell_str)
            merged = " ".join(parts)
            # Replace multiple spaces with a single space
            self.merged_headers.append(" ".join(merged.split()))

        logger.info(f"Header rows detected: {self.header_rows}")
        logger.info(f"Merged headers count: {len(self.merged_headers)}")

    def _detect_company(self):
        """Extract the company name from the rows above headers and find the API fragment."""
        title_text = ""
        # Check the first few rows (prior to header rows)
        first_header_row = self.header_rows[0]
        for r_idx in range(first_header_row):
            for col_idx in range(len(self.df.columns)):
                cell_val = clean_val(self.df.iloc[r_idx, col_idx])
                if len(cell_val) > 10 and any(keyword in cell_val.lower() for keyword in ["private", "pvt", "limited", "ltd", "solutions", "software", "informatics"]):
                    title_text = cell_val
                    break
            if title_text:
                break

        # Fallback to file name if no title found in sheet
        if not title_text:
            title_text = os.path.basename(self.file_path)

        title_lower = title_text.lower()
        logger.info(f"Extracted sheet title text: '{title_text}'")

        # Map to known API company fragments
        if "originator" in title_lower:
            self.company_fragment = "Originator"
            self.display_company = "Originator Informatics Pvt. Ltd."
            self.display_branch = "Varanasi (TEK)"
        elif "code crewz" in title_lower:
            self.company_fragment = "Code Crewz"
            self.display_company = "Code Crewzs"
            self.display_branch = "Varanasi"
        elif "infoserv" in title_lower:
            self.company_fragment = "Infoserv"
            self.display_company = "Infoserv"
            self.display_branch = "Varanasi"
        elif "init" in title_lower:
            self.company_fragment = "INIT"
            self.display_company = "INIT"
            self.display_branch = "Varanasi"
        elif "jobvritta" in title_lower:
            self.company_fragment = "Jobvritta"
            self.display_company = "Jobvritta"
            self.display_branch = "Varanasi"
        elif "tek inspirations" in title_lower:
            self.company_fragment = "Tek Inspirations"
            self.display_company = "Tek Inspirations Pvt. Ltd."
            self.display_branch = "Varanasi (TEK)"
        elif "legit guru" in title_lower:
            self.company_fragment = "Legit Guru"
            self.display_company = "Legit Guru"
            self.display_branch = "Agra"
        elif "adventa tech software" in title_lower:
            self.company_fragment = "Adventa Tech Software Solutions"
            self.display_company = "Adventa Tech Software Solutions Pvt Ltd"
            self.display_branch = "Bhubaneswar"
        elif "adventa tech" in title_lower or "adventa" in title_lower:
            # Check sheet name or title to see if it is the Bhubaneswar branch or Agra
            if "bhubaneswar" in title_lower or self.sheet_name.lower() == "adventa":
                self.company_fragment = "Adventa Tech Software Solutions"
                self.display_company = "Adventa Tech Software Solutions Pvt Ltd"
                self.display_branch = "Bhubaneswar"
            else:
                self.company_fragment = "Adventa Tech Pvt. Ltd."
                self.display_company = "Adventa Tech Pvt. Ltd."
                self.display_branch = "Agra"
        elif "vyze" in title_lower:
            self.company_fragment = "Vyze"
            self.display_company = "Vyze"
            self.display_branch = "Bhubaneswar"
        elif "infusive" in title_lower:
            self.company_fragment = "Infusive"
            self.display_company = "Infusive"
            # Deduce branch from sheet name (Agra, Jaipur, Noida)
            if "jaipur" in self.sheet_name.lower():
                self.display_branch = "Jaipur"
            elif "noida" in self.sheet_name.lower():
                self.display_branch = "Noida"
            else:
                self.display_branch = "Agra"
        else:
            # Ultimate generic fallback - extract first two words of title
            words = [w for w in re.split(r'\W+', title_text) if w]
            self.company_fragment = " ".join(words[:2]) if len(words) >= 2 else title_text
            self.display_company = title_text

        logger.info(f"Auto-detected Company fragment: '{self.company_fragment}' (Branch: '{self.display_branch}')")

    def _detect_core_columns(self):
        """Locate Employee Name, Employee Code, and Net Salary columns dynamically."""
        for idx, header in enumerate(self.merged_headers):
            norm = normalize_header_text(header)
            
            # 1. Employee Name column
            if "name" in norm and not any(k in norm for k in ["father", "company", "bank", "branch"]):
                self.employee_name_col = idx
            
            # 2. Employee Code column
            if any(k in norm for k in ["code", "emp id", "employee id", "hrlenseid"]) and not "company" in norm:
                # If there's another employee code column, we prefer the one that is cleaner or has higher index
                self.employee_code_col = idx
            elif norm == "emp" and self.employee_code_col == -1:
                # Fallback for solitary "Emp" header
                self.employee_code_col = idx

            # 3. Net Salary column
            if any(k in norm for k in ["net payable", "net salary", "net amount", "total net", "net payable salary"]) or (norm == "payable" and self.net_salary_col == -1):
                self.net_salary_col = idx

        # Also detect HrLense ID column explicitly
        for idx, header in enumerate(self.merged_headers):
            norm = normalize_header_text(header)
            if "hrlense" in norm or "hr lense" in norm:
                self.hrlense_id_col = idx
                break

        # Fallback if employee code column not found
        if self.employee_code_col == -1:
            for idx, header in enumerate(self.merged_headers):
                norm = normalize_header_text(header)
                if norm == "emp":
                    self.employee_code_col = idx
                    break

        # Let's log core findings
        logger.info(f"Core column detection: Name Col={self.employee_name_col}, Code Col={self.employee_code_col}, HrLense ID Col={self.hrlense_id_col}, Net Salary Col={self.net_salary_col}")

    def _map_logical_columns(self):
        """Map logical payroll fields to Excel column indices using clean headers. Handles duplicates (prefers latter index)."""
        self.col_mapping = {}
        
        # Override specifically for Infusive Pvt. Ltd. branches to achieve extremely strict, correct column mapping
        if self.company_fragment == "Infusive":
            for idx, header in enumerate(self.merged_headers):
                norm = normalize_header_text(header)
                
                if norm == "emp name":
                    self.employee_name_col = idx
                    self.col_mapping["employeeName"] = (idx, header)
                elif norm == "emp code":
                    self.employee_code_col = idx
                    self.col_mapping["employeeCode"] = (idx, header)
                elif "hrlense id" in norm:
                    self.hrlense_id_col = idx
                elif norm in ["net payable", "net payable amt"]:
                    self.net_salary_col = idx
                    self.col_mapping["netSalary"] = (idx, header)
                elif norm == "total payable days":
                    self.col_mapping["paidDays"] = (idx, header)
                elif norm == "deductions pf":
                    self.col_mapping["employee_PF"] = (idx, header)
                elif norm == "esi":
                    self.col_mapping["esic_Employee"] = (idx, header)
                elif norm == "tds":
                    self.col_mapping["tds"] = (idx, header)
                elif norm == "cl mar-26":
                    self.col_mapping["balance_Leave"] = (idx, header)
            
            # The actual earned components are located after the 'total gross amt.' column (Col 17 in Agra, Col 25 in Jaipur/Noida)
            gross_idx = -1
            for idx, header in enumerate(self.merged_headers):
                norm = normalize_header_text(header)
                if norm == "total gross amt":
                    gross_idx = idx
                    break
            if gross_idx == -1:
                gross_idx = 16
                
            for idx in range(gross_idx, len(self.merged_headers)):
                header = self.merged_headers[idx]
                norm = normalize_header_text(header)
                
                if norm == "basic + da amt":
                    self.col_mapping["basic"] = (idx, header)
                elif norm == "hra amt":
                    self.col_mapping["hra"] = (idx, header)
                elif norm == "conv amt":
                    self.col_mapping["conveyance_Allowance"] = (idx, header)
                elif norm == "special allowance":
                    self.col_mapping["incentive"] = (idx, header)
            
            # Ensure employee name and code are explicitly mapped if detected
            if self.employee_name_col != -1 and "employeeName" not in self.col_mapping:
                self.col_mapping["employeeName"] = (self.employee_name_col, self.merged_headers[self.employee_name_col])
            if self.employee_code_col != -1 and "employeeCode" not in self.col_mapping:
                self.col_mapping["employeeCode"] = (self.employee_code_col, self.merged_headers[self.employee_code_col])
            if self.net_salary_col != -1 and "netSalary" not in self.col_mapping:
                self.col_mapping["netSalary"] = (self.net_salary_col, self.merged_headers[self.net_salary_col])
                
            # Logging mappings
            for k, v in self.col_mapping.items():
                logger.info(f"Mapped Field: {k:<25} -> Col Index {v[0]:<3} ({v[1]})")
            return

        # We loop through target fields and match them to the merged headers
        for field, keywords in TARGET_FIELDS.items():
            best_col = -1
            best_header = ""
            
            # Check all column headers
            for idx, header in enumerate(self.merged_headers):
                norm = normalize_header_text(header)
                
                # Check for keyword matches
                match_found = False
                for kw in keywords:
                    norm_kw = normalize_header_text(kw)
                    if norm_kw == norm or norm_kw in norm:
                        match_found = True
                        break
                
                if match_found:
                    # Guard against matching deductions or other fields as earnings
                    if field in ["basic", "hra", "conveyance_Allowance", "incentive"]:
                        # For earnings, make sure it does not contain 'fixed' or 'standard' or 'actual' if we have duplicates
                        # Actually, simply preferring the LARGEST column index handles Fixed vs Earned duplicates perfectly!
                        best_col = idx
                        best_header = header
                    elif field in ["employee_PF", "esic_Employee", "tds"]:
                        # Deductions
                        best_col = idx
                        best_header = header
                    else:
                        # Other fields like netSalary, paidDays, balance_Leave
                        best_col = idx
                        best_header = header
 
            if best_col != -1:
                self.col_mapping[field] = (best_col, best_header)

        # Ensure employee name and code are explicitly mapped if detected
        if self.employee_name_col != -1:
            self.col_mapping["employeeName"] = (self.employee_name_col, self.merged_headers[self.employee_name_col])
        if self.employee_code_col != -1:
            self.col_mapping["employeeCode"] = (self.employee_code_col, self.merged_headers[self.employee_code_col])
        if self.net_salary_col != -1:
            self.col_mapping["netSalary"] = (self.net_salary_col, self.merged_headers[self.net_salary_col])

        # Logging mappings
        for k, v in self.col_mapping.items():
            logger.info(f"Mapped Field: {k:<25} -> Col Index {v[0]:<3} ({v[1]})")

    def _validate_structure(self):
        """Fail fast if expected core headers are missing."""
        self.header_validation = {
            "has_header_rows": len(self.header_rows) > 0,
            "has_employee_name": self.employee_name_col != -1,
            "has_employee_code": self.employee_code_col != -1,
            "has_net_salary": self.net_salary_col != -1,
            "column_count": len(self.df.columns)
        }
        
        if not self.header_validation["has_header_rows"]:
            self.status = "FAIL"
            raise StructuralValidationError("Could not detect any valid header rows.")
            
        if not self.header_validation["has_employee_name"]:
            self.status = "FAIL"
            raise StructuralValidationError("Employee Name column not found in headers.")
            
        if not self.header_validation["has_employee_code"]:
            self.status = "FAIL"
            raise StructuralValidationError("Employee Code column not found in headers.")
            
        if not self.header_validation["has_net_salary"]:
            self.status = "FAIL"
            raise StructuralValidationError("Net Payable column not found in headers.")

    def parse_data(self) -> list:
        """Step 3: Parse and validate excel records, skipping headers, blank rows, and subtotals."""
        records = []
        start_row = self.header_rows[-1] + 1
        consecutive_blanks = 0
        seen_codes = set()
        seen_names = set()
        
        total_rows_parsed = 0
        
        for r_idx in range(start_row, len(self.df)):
            row = self.df.iloc[r_idx]
            
            # Check if row is completely blank
            is_row_blank = all(pd.isna(val) or str(val).strip() == "" for val in row)
            
            if is_row_blank:
                consecutive_blanks += 1
                if consecutive_blanks >= 3:
                    logger.info(f"Stopped parsing after {consecutive_blanks} consecutive blank rows at row {r_idx}")
                    break
                continue
                
            consecutive_blanks = 0
            
            # Extract name and code
            name_val = row.iloc[self.employee_name_col] if self.employee_name_col < len(row) else None
            code_val = row.iloc[self.employee_code_col] if self.employee_code_col < len(row) else None
            
            # Fallback to HrLense ID if code is blank (especially important for Infusive branches)
            if not clean_val(code_val) and hasattr(self, "hrlense_id_col") and self.hrlense_id_col != -1:
                fallback_code = row.iloc[self.hrlense_id_col] if self.hrlense_id_col < len(row) else None
                if clean_val(fallback_code):
                    code_val = fallback_code
            
            # Detect subtotal/footer rows
            name_str = clean_val(name_val).lower()
            code_str = clean_val(code_val).lower()
            
            is_subtotal = any(x in name_str or x in code_str for x in ["total", "subtotal", "grand total", "dept total", "amount", "summary", "total:"])
            
            # If name/code is empty but net salary is present, it might be a footer row
            if not name_str and not code_str:
                net_val = row.iloc[self.net_salary_col] if self.net_salary_col < len(row) else None
                if clean_val(net_val):
                    is_subtotal = True
                    
            if is_subtotal:
                logger.info(f"Skipping subtotal/footer row {r_idx}: Name='{name_val}', Code='{code_val}'")
                continue

            # Skip header lookalikes or other metadata rows (e.g. serial numbers without names/codes)
            if not name_str and not code_str:
                continue

            total_rows_parsed += 1

            # Validate Duplicates & Mandatory IDs
            emp_code = normalize_code(code_val)
            emp_name = normalize_name(name_val)
            
            row_errors = []
            
            if not emp_code:
                row_errors.append("Blank Employee ID")
                self.status = "WARNING"
                self.duplicates.append({
                    "type": "Blank ID",
                    "row": r_idx + 1,
                    "employee_name": str(name_val).strip()
                })
            else:
                if emp_code in seen_codes:
                    row_errors.append(f"Duplicate Employee Code: {emp_code}")
                    self.status = "WARNING"
                    self.duplicates.append({
                        "type": "Duplicate Code",
                        "row": r_idx + 1,
                        "value": emp_code,
                        "employee_name": str(name_val).strip()
                    })
                seen_codes.add(emp_code)

            if emp_name:
                if emp_name in seen_names:
                    self.duplicates.append({
                        "type": "Duplicate Name",
                        "row": r_idx + 1,
                        "value": str(name_val).strip(),
                        "employee_code": emp_code
                    })
                seen_names.add(emp_name)

            # Build record dict
            rec = {
                "_row": r_idx + 1,
                "employeeName": str(name_val).strip(),
                "employeeCode": emp_code,
                "errors": row_errors
            }
            
            # Parse mapped logical fields
            for field, (col_idx, _) in self.col_mapping.items():
                if field in ["employeeName", "employeeCode"]:
                    continue
                val = row.iloc[col_idx] if col_idx < len(row) else None
                if field == "balance_Leave":
                    rec[field] = clean_val(val) # Preserve leave balance as string/val
                else:
                    rec[field] = safe_float(val)

            # Numeric Consistency Check (earnings - deductions == netSalary)
            basic_val = rec.get("basic", 0.0)
            hra_val = rec.get("hra", 0.0)
            conv_val = rec.get("conveyance_Allowance", 0.0)
            inc_val = rec.get("incentive", 0.0)
            pf_val = rec.get("employee_PF", 0.0)
            esi_val = rec.get("esic_Employee", 0.0)
            tds_val = rec.get("tds", 0.0)
            net_val = rec.get("netSalary", 0.0)
            
            calculated_earnings = round(basic_val + hra_val + conv_val + inc_val, 2)
            calculated_deductions = round(pf_val + esi_val + tds_val, 2)
            calculated_net = round(calculated_earnings - calculated_deductions, 2)
            
            # If netSalary column is mapped, verify consistency
            if "netSalary" in self.col_mapping and abs(calculated_net - net_val) > 1.0:
                logger.warning(f"Row {r_idx+1} numeric inconsistency: Calculated Net={calculated_net}, Excel Net={net_val} (diff={round(calculated_net - net_val, 2)})")
                
            records.append(rec)

        logger.info(f"Successfully parsed {len(records)} employee records.")
        return records

    def compare_against_api(self, excel_records: list, api_records: list) -> dict:
        """Step 4: Perform exact employee code matching and normalized name matching with API records."""
        # Build API lookups
        composite_lookup = {}
        name_lookup = {}
        for rec in api_records:
            code = normalize_code(rec.get("employeeCode", ""))
            name = normalize_name(rec.get("employeeName", ""))
            if code and name:
                composite_lookup[(name, code)] = rec
            if code:
                composite_lookup[("", code)] = rec
            if name:
                name_lookup[name] = rec

        matched_records = []
        self.mismatches = []
        missing_in_api = []
        matched_api_keys = set()

        for er in excel_records:
            excel_name = er["employeeName"]
            excel_code = er["employeeCode"]
            
            norm_name = normalize_name(excel_name)
            norm_code = normalize_code(excel_code)
            
            # Try matching by ID first (exact)
            api_rec = None
            match_type = ""
            
            if norm_code:
                # Lookup by code
                api_rec = composite_lookup.get(("", norm_code))
                if api_rec:
                    match_type = "EXACT_CODE"
            
            if not api_rec and norm_name:
                # Fallback to name match
                api_rec = name_lookup.get(norm_name)
                if api_rec:
                    match_type = "EXACT_NAME"
                else:
                    # Fuzzy match with name
                    matches = difflib.get_close_matches(norm_name, list(name_lookup.keys()), n=1, cutoff=0.85)
                    if matches:
                        api_rec = name_lookup[matches[0]]
                        match_type = f"FUZZY_NAME({matches[0]})"

            if not api_rec:
                # Missing in API
                self.status = "WARNING"
                missing_in_api.append({
                    "employee_code": excel_code,
                    "employee_name": excel_name,
                    "excel_row": er["_row"],
                    "net_payable": er.get("netSalary", 0.0)
                })
                continue

            # Record match
            api_code = normalize_code(api_rec.get("employeeCode", ""))
            api_name = normalize_name(api_rec.get("employeeName", ""))
            matched_api_keys.add((api_name, api_code))
            matched_api_keys.add(("", api_code))
            
            matched_records.append({
                "employee_code": excel_code,
                "employee_name": excel_name,
                "excel_row": er["_row"],
                "match_type": match_type
            })

            # Compare all mapped payroll fields
            for logical_field in self.col_mapping.keys():
                if logical_field in ["employeeName", "employeeCode"]:
                    continue
                
                excel_val = er.get(logical_field)
                api_field = logical_field
                
                # Special handling for field names if they differ in API response
                api_val = api_rec.get(api_field)
                
                if logical_field == "balance_Leave":
                    # Compare leave balance (handle as floats if numeric, else normalized strings)
                    excel_num = safe_float(excel_val)
                    api_num = safe_float(api_val)
                    
                    # API leave balance represents current status (May), while Excel holds balance as of April.
                    # The accrual difference is typically +1.5, +1.0, or 0.0 depending on branch/policy.
                    is_match = (
                        abs(api_num - excel_num) <= 0.05 or 
                        abs(api_num - (excel_num + 1.5)) <= 0.05 or 
                        abs(api_num - (excel_num + 1.0)) <= 0.05
                    )
                    
                    if not is_match:
                        self.status = "FAIL"
                        self.mismatches.append({
                            "employee_code": excel_code,
                            "employee_name": excel_name,
                            "excel_row": er["_row"],
                            "field": logical_field,
                            "expected_value": api_val if api_val is not None else "—",
                            "actual_value": excel_val if excel_val is not None else "—",
                            "source_column": self.col_mapping[logical_field][1],
                            "remarks": f"Leave Balance mismatch (Excel: {excel_val}, API: {api_val}; does not match standard accrual)"
                        })
                else:
                    # Compare numeric fields
                    excel_num = safe_float(excel_val)
                    api_num = safe_float(api_val)
                    if excel_num != api_num:
                        self.status = "FAIL"
                        self.mismatches.append({
                            "employee_code": excel_code,
                            "employee_name": excel_name,
                            "excel_row": er["_row"],
                            "field": logical_field,
                            "expected_value": api_num,
                            "actual_value": excel_num,
                            "source_column": self.col_mapping[logical_field][1],
                            "remarks": f"Numeric value mismatch for {logical_field}"
                        })

        # Identify missing in Excel (exists in API but not spreadsheet)
        missing_in_excel = []
        for rec in api_records:
            code = normalize_code(rec.get("employeeCode", ""))
            name = normalize_name(rec.get("employeeName", ""))
            if (name, code) not in matched_api_keys and ("", code) not in matched_api_keys:
                missing_in_excel.append({
                    "employee_code": code,
                    "employee_name": rec.get("employeeName", ""),
                    "net_payable": safe_float(rec.get("netSalary", 0.0))
                })

        # Parser confidence score
        total_excel_rows = len(excel_records)
        matched_count = len(matched_records)
        confidence_score = round((matched_count / total_excel_rows) * 100, 2) if total_excel_rows > 0 else 0.0

        return {
            "status": self.status,
            "matched": matched_records,
            "mismatches": self.mismatches,
            "missing_in_api": missing_in_api,
            "missing_in_payroll": missing_in_excel,
            "confidence_score": confidence_score
        }

    def generate_json_report(self, comparison_results: dict, output_dir: str = "reports/payroll json report") -> str:
        """Step 5: Generate the deterministic JSON report file, backing up previous if existing."""
        os.makedirs(output_dir, exist_ok=True)
        filename = f"reconciliation_report_{re.sub(r'[^a-zA-Z0-9]', '_', self.display_company)[:30]}_{self.sheet_name}.json"
        report_path = os.path.join(output_dir, filename)

        # Back up existing report if it exists
        if os.path.exists(report_path):
            backup_path = f"{report_path}.bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            shutil.copy2(report_path, backup_path)
            logger.info(f"Backed up existing JSON report to: {backup_path}")

        # Construct JSON output structure
        report_data = {
            "file": os.path.basename(self.file_path),
            "sheet": self.sheet_name,
            "status": self.status,
            "header_validation": self.header_validation,
            "mapping": {k: {"col_index": v[0], "header": v[1]} for k, v in self.col_mapping.items()},
            "duplicates": self.duplicates,
            "mismatches": self.mismatches,
            "summary": {
                "total_rows": len(comparison_results["matched"]) + len(comparison_results["missing_in_api"]),
                "matched": len(comparison_results["matched"]),
                "mismatched": len(comparison_results["mismatches"]),
                "missing_in_api": len(comparison_results["missing_in_api"]),
                "missing_in_payroll": len(comparison_results["missing_in_payroll"]),
                "confidence_score": comparison_results["confidence_score"]
            }
        }

        with open(report_path, "w") as f:
            json.dump(report_data, f, indent=2)

        logger.info(f"JSON report written: {report_path}")
        return report_path

    def generate_excel_report(self, excel_records: list, api_records: list, comparison_results: dict, output_dir: str = "reports/payroll data report") -> str:
        """Step 5: Create a beautiful, detailed layout-independent Excel validation report workbook."""
        os.makedirs(output_dir, exist_ok=True)
        slug = re.sub(r"[^a-zA-Z0-9]", "_", self.display_company)[:30]
        filename = f"payroll_validation_{slug}_{self.sheet_name}.xlsx"
        report_path = os.path.join(output_dir, filename)

        # Back up existing Excel report if it exists
        if os.path.exists(report_path):
            backup_path = f"{report_path}.bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            try:
                shutil.copy2(report_path, backup_path)
                logger.info(f"Backed up existing Excel report to: {backup_path}")
            except Exception as e:
                logger.warning(f"Could not back up existing Excel report: {e}")

        wb = openpyxl.Workbook()
        wb.remove(wb.active) # remove default sheet

        MONTH_NAME = "APRIL 2026"
        TITLE = f"PAYROLL VALIDATION REPORT  —  {self.display_company}  —  {self.display_branch}  —  {MONTH_NAME}"

        HDR_FILL = PatternFill("solid", fgColor="1F4E79")
        HDR_FONT = Font(bold=True, color="FFFFFF", size=10) # Condensed header font
        ROW_FONT = Font(size=9) # Condensed row font
        BOLD_ROW_FONT = Font(bold=True, size=9)
        SEP_FILL = PatternFill("solid", fgColor="D9E1F2")
        FAIL_FILL = PatternFill("solid", fgColor="FFE0E0")
        PASS_FILL = PatternFill("solid", fgColor="E2EFDA")
        WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
        
        thin = Side(style="thin")
        bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

        def make_sheet(title):
            ws = wb.create_sheet(title=title)
            ws.append([TITLE])
            ws["A1"].font = Font(bold=True, size=11) # Condensed title
            ws.row_dimensions[1].height = 20
            ws.append([])
            ws.row_dimensions[2].height = 8
            return ws

        def write_header(ws, cols):
            ws.append(cols)
            row = ws.max_row
            ws.row_dimensions[row].height = 22 # Condensed header height
            for c in range(1, len(cols) + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill, cell.font = HDR_FILL, HDR_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = bdr

        def write_row(ws, values, fill=None):
            ws.append(values)
            row = ws.max_row
            ws.row_dimensions[row].height = 16 # Condensed row height
            for c in range(1, len(values) + 1):
                cell = ws.cell(row=row, column=c)
                cell.border = bdr
                cell.font = ROW_FONT
                
                # Check alignment based on value type
                val = values[c - 1]
                if isinstance(val, (int, float)):
                    align = Alignment(horizontal="right", vertical="center", wrap_text=True)
                else:
                    align = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.alignment = align
                
                if fill:
                    cell.fill = fill

        def write_total(ws, count):
            ws.append([])
            ws.row_dimensions[ws.max_row].height = 8
            ws.append(["Total:", count])
            row = ws.max_row
            ws.row_dimensions[row].height = 16
            c1 = ws.cell(row=row, column=1)
            c2 = ws.cell(row=row, column=2)
            c1.font = BOLD_ROW_FONT
            c2.font = BOLD_ROW_FONT
            c1.border = bdr
            c2.border = bdr

        def autofit(ws):
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                
                max_len = 0
                for cell in col:
                    if cell.row == 1: # Skip title row length
                        continue
                    val_str = str(cell.value or "")
                    lines = val_str.split("\n")
                    for line in lines:
                        if len(line) > max_len:
                            max_len = len(line)
                
                # Scale width between 10 and 50 based on content length
                w = min(max(max_len + 4, 10), 50)
                
                # Keep serial number column narrow
                header_val = str(col[2].value or "").lower() if len(col) > 2 else ""
                if col_letter == "A" or header_val == "#":
                    w = min(max(max_len + 3, 5), 8)
                    
                ws.column_dimensions[col_letter].width = w

        # ----------------------------------------------------
        # T0 FIELD DETAIL (Wide View of all matched records and fields)
        # ----------------------------------------------------
        ws0 = make_sheet("T0 Field Detail")
        ws0.append(["TABLE 0: FIELD-LEVEL DETAIL (Wide Comparison)"])
        ws0[f"A{ws0.max_row}"].font = Font(bold=True, size=11)

        # Compile detail columns
        detail_fields = [k for k in self.col_mapping.keys() if k not in ["employeeName", "employeeCode"]]
        field_cols = []
        for df_key in detail_fields:
            lbl = df_key.replace("_", " ").title()
            field_cols.extend([f"{lbl} Excel", f"{lbl} API", f"{lbl} Diff"])

        all_headers = ["#", "Employee Name", "Employee Code", "Match Type"] + field_cols + ["Final Status"]
        write_header(ws0, all_headers)

        fs_col = len(all_headers)
        
        # Build wide rows
        t0_idx = 1
        for er in excel_records:
            excel_code = er["employeeCode"]
            excel_name = er["employeeName"]
            
            # Find in matched
            match_info = next((m for m in comparison_results["matched"] if m["employee_code"] == excel_code), None)
            if not match_info:
                continue
                
            row_mismatches = [m for m in comparison_results["mismatches"] if m["employee_code"] == excel_code]
            status_text = "Fail" if row_mismatches else "Pass"
            row_fill = FAIL_FILL if status_text == "Fail" else PASS_FILL

            vals = [t0_idx, excel_name, excel_code, match_info["match_type"]]
            t0_idx += 1
            
            # Match against API record
            api_rec = None
            for rec in api_records:
                if normalize_code(rec.get("employeeCode", "")) == excel_code:
                    api_rec = rec
                    break

            for df_key in detail_fields:
                excel_v = er.get(df_key, 0.0)
                api_v = 0.0
                if api_rec:
                    api_v = safe_float(api_rec.get(df_key))
                
                if df_key == "balance_Leave":
                    # Textual / leave balance handling
                    excel_v_str = str(excel_v)
                    api_v_str = str(api_rec.get(df_key) if api_rec else "—")
                    diff_str = "—" if excel_v_str == api_v_str else "Mismatch"
                    vals.extend([excel_v_str, api_v_str, diff_str])
                else:
                    diff_num = round(excel_v - api_v, 2)
                    vals.extend([excel_v, api_v, diff_num])

            vals.append(status_text)
            write_row(ws0, vals, row_fill)

        autofit(ws0)

        # ----------------------------------------------------
        # T1 SUMMARY (Exact Original Structure)
        # ----------------------------------------------------
        ws1 = make_sheet("T1 Summary")
        ws1.append(["TABLE 1: RECONCILIATION SUMMARY"])
        ws1[f"A{ws1.max_row}"].font = Font(bold=True, size=12)
        write_header(ws1, ["Category", "Count"])
        
        # Calculate exact metrics
        n_matched = len(comparison_results["matched"])
        n_missing_in_api = len(comparison_results["missing_in_api"])
        n_missing_in_payroll = len(comparison_results["missing_in_payroll"])
        
        n_exact = sum(1 for m in comparison_results["matched"] if "FUZZY" not in m["match_type"])
        n_fuzzy = sum(1 for m in comparison_results["matched"] if "FUZZY" in m["match_type"])
        
        # Mismatches on Net Payable
        net_salary_mismatched_codes = {m["employee_code"] for m in comparison_results["mismatches"] if m["field"] == "netSalary"}
        n_salary_match = sum(1 for m in comparison_results["matched"] if m["employee_code"] not in net_salary_mismatched_codes)
        n_salary_diff = sum(1 for m in comparison_results["matched"] if m["employee_code"] in net_salary_mismatched_codes)
        
        t1_rows = [
            ("Excel Records (Source)", n_matched + n_missing_in_api),
            ("API Records (Payroll)", len(api_records)),
            ("---", "---"),
            ("Matched (Exact)", n_exact),
            ("Total Matched", n_matched),
            ("Matched (Fuzzy / Spelling)", n_fuzzy),
            ("---", "---"),
            ("Salary Match (= Paid)", n_salary_match),
            ("Salary Diff (!= Paid)", n_salary_diff),
            ("---", "---"),
            ("Missing in API", n_missing_in_api),
            ("Missing in Excel", n_missing_in_payroll),
            ("---", "---"),
            ("Parser Confidence Score (%)", f"{comparison_results['confidence_score']}%")
        ]
        
        for cat, cnt in t1_rows:
            is_sep = cat == "---"
            write_row(ws1, [cat, cnt], SEP_FILL if is_sep else None)
            if cat.startswith(("Matched", "Unmatched", "Parser", "Excel", "API")):
                ws1.cell(row=ws1.max_row, column=1).font = Font(bold=True)
                ws1.cell(row=ws1.max_row, column=2).font = Font(bold=True)
        autofit(ws1)

        # ----------------------------------------------------
        # T2 DIFFERENCE TABLE (Wide Layout - One Employee Per Row)
        # ----------------------------------------------------
        ws2 = make_sheet("T2 Differences")
        ws2.append(["TABLE 2: DIFFERENCE RECONCILIATION (Wide View)"])
        ws2[f"A{ws2.max_row}"].font = Font(bold=True, size=12)
        
        # Compile detail columns
        detail_fields = [k for k in self.col_mapping.keys() if k not in ["employeeName", "employeeCode"]]
        field_cols = []
        for df_key in detail_fields:
            lbl = df_key.replace("_", " ").title()
            field_cols.extend([f"{lbl} Excel", f"{lbl} API", f"{lbl} Diff"])

        all_headers_t2 = ["#", "Employee Name", "Employee Code", "Match Type"] + field_cols + ["Mismatched Fields"]
        write_header(ws2, all_headers_t2)

        # Get codes that have mismatches
        mismatched_codes = {m["employee_code"] for m in comparison_results["mismatches"]}
        
        t2_idx = 1
        for er in excel_records:
            excel_code = er["employeeCode"]
            excel_name = er["employeeName"]
            
            if excel_code not in mismatched_codes:
                continue

            match_info = next((m for m in comparison_results["matched"] if m["employee_code"] == excel_code), None)
            match_type = match_info["match_type"] if match_info else "—"

            row_mismatches = [m for m in comparison_results["mismatches"] if m["employee_code"] == excel_code]
            mismatched_field_names = [m["field"].replace("_", " ").title() for m in row_mismatches]
            mismatched_str = ", ".join(mismatched_field_names)

            vals = [t2_idx, excel_name, excel_code, match_type]
            t2_idx += 1

            api_rec = None
            for rec in api_records:
                if normalize_code(rec.get("employeeCode", "")) == excel_code:
                    api_rec = rec
                    break

            for df_key in detail_fields:
                excel_v = er.get(df_key, 0.0)
                api_v = 0.0
                if api_rec:
                    api_v = safe_float(api_rec.get(df_key))
                
                if df_key == "balance_Leave":
                    # Textual / leave balance handling
                    excel_v_str = str(excel_v)
                    api_v_str = str(api_rec.get(df_key) if api_rec else "—")
                    diff_str = "—" if excel_v_str == api_v_str else "Mismatch"
                    vals.extend([excel_v_str, api_v_str, diff_str])
                else:
                    diff_num = round(excel_v - api_v, 2)
                    vals.extend([excel_v, api_v, diff_num])

            vals.append(mismatched_str)
            write_row(ws2, vals)
            
            # Now highlight the specific cells that had mismatches!
            row_num = ws2.max_row
            mismatch_field_keys = {m["field"] for m in row_mismatches}
            
            # Highlight Employee Name and Code
            ws2.cell(row=row_num, column=2).fill = WARN_FILL
            ws2.cell(row=row_num, column=3).fill = WARN_FILL
            
            # Highlight each mismatched field's Excel, API, and Diff cells in red
            for f_idx, df_key in enumerate(detail_fields):
                if df_key in mismatch_field_keys:
                    start_col = 5 + f_idx * 3 # 1-indexed: # is 1, Name is 2, Code is 3, MatchType is 4, so first field Excel is at col 5
                    for col_offset in range(3):
                        cell = ws2.cell(row=row_num, column=start_col + col_offset)
                        cell.fill = FAIL_FILL

        write_total(ws2, t2_idx - 1)
        autofit(ws2)

        # ----------------------------------------------------
        # T3 DUPLICATES / DATA WARNINGS
        # ----------------------------------------------------
        ws3 = make_sheet("T3 Duplicates")
        ws3.append(["TABLE 3: DUPLICATE AND DATA VALIDATION ALERTS"])
        ws3[f"A{ws3.max_row}"].font = Font(bold=True, size=12)
        write_header(ws3, ["#", "Alert Type", "Excel Row", "Value", "Employee Info"])
        
        for idx, d in enumerate(self.duplicates, 1):
            emp_info = d.get("employee_name", d.get("employee_code", ""))
            write_row(ws3, [idx, d["type"], d["row"], d.get("value", "—"), emp_info], WARN_FILL)
        write_total(ws3, len(self.duplicates))
        autofit(ws3)

        # ----------------------------------------------------
        # T4 UNMATCHED IN API (Missing in API)
        # ----------------------------------------------------
        ws4 = make_sheet("T4 Missing in API")
        ws4.append(["TABLE 4: UNMATCHED EMPLOYEES (In Excel, No Match in API)"])
        ws4[f"A{ws4.max_row}"].font = Font(bold=True, size=12)
        write_header(ws4, ["#", "Employee Name", "Employee Code", "Excel Row", "Net Payable Salary"])
        
        for idx, u in enumerate(comparison_results["missing_in_api"], 1):
            write_row(ws4, [idx, u["employee_name"], u["employee_code"], u["excel_row"], u["net_payable"]], WARN_FILL)
        write_total(ws4, len(comparison_results["missing_in_api"]))
        autofit(ws4)

        # ----------------------------------------------------
        # T5 UNMATCHED IN EXCEL (Missing in Excel)
        # ----------------------------------------------------
        ws5 = make_sheet("T5 Missing in Excel")
        ws5.append(["TABLE 5: UNMATCHED EMPLOYEES (In API, Not found in Excel)"])
        ws5[f"A{ws5.max_row}"].font = Font(bold=True, size=12)
        write_header(ws5, ["#", "Employee Name (API)", "Employee Code (API)", "Expected Net Payable"])
        
        for idx, u in enumerate(comparison_results["missing_in_payroll"], 1):
            write_row(ws5, [idx, u["employee_name"], u["employee_code"], u["net_payable"]], WARN_FILL)
        write_total(ws5, len(comparison_results["missing_in_payroll"]))
        autofit(ws5)

        wb.save(report_path)
        logger.info(f"Excel report written: {report_path}")
        return report_path
