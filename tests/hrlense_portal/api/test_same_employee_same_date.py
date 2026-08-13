"""
Absence Case Analysis: Same Employee on the Exact Same Date.
Generates: 'reports/Same_Employee_Same_Date.xlsx' and 'reports/Same_Employee_Same_Date.md'
Columns:
- Employee ID
- Employee Name
- Company
- Branch
- Absence Date
- Case ID
- Case Status
- Continuous Absent Days
- Threshold Days
- Created Date/Time
"""

import os
import json
import logging
import pytest
import pandas as pd
from collections import defaultdict
from core.config import settings
from utils.api.base_api import get

logger = logging.getLogger(__name__)


def fetch_all_absence_cases() -> list[dict]:
    """Fetches complete absence cases dataset via API."""
    all_records = []
    first = 0
    rows = 500
    while True:
        lazy_params = json.dumps({
            "first": first,
            "rows": rows,
            "page": first // rows,
            "sortField": "id",
            "sortOrder": -1
        })
        try:
            resp = get("AbsenceCase/list", params={"lazyParams": lazy_params, "search": "", "status": ""})
            records = []
            if isinstance(resp, list):
                records = resp
            elif isinstance(resp, dict):
                if "data" in resp:
                    records = resp["data"] if isinstance(resp["data"], list) else resp["data"].get("rows", [])
                elif "rows" in resp:
                    records = resp["rows"]
            if not records:
                break
            all_records.extend(records)
            if len(records) < rows or len(all_records) >= 2500:
                break
            first += rows
        except Exception:
            break
    return all_records


@pytest.mark.api
def test_generate_same_employee_same_date_excel():
    """
    Identifies all instances where the SAME Employee has MULTIPLE Absence Case IDs
    logged on the EXACT SAME Date.
    Generates 'reports/Same_Employee_Same_Date.xlsx'.
    """
    logger.info("Fetching Absence Cases for Same-Employee Same-Date collision audit...")
    cases = fetch_all_absence_cases()
    assert len(cases) > 0, "No records retrieved from API."

    # Group by (Employee ID, Absence Date)
    emp_date_groups = defaultdict(list)

    for c in cases:
        emp_id = str(c.get("employee_Id") or c.get("employee_Code") or c.get("employeeId") or "N/A")
        emp_name = str(c.get("employee_Name") or c.get("employeeName") or "Unknown").strip()
        raw_date = str(c.get("start_Date") or c.get("startDate") or "")
        absence_date = raw_date.split("T")[0] if "T" in raw_date else raw_date[:10]
        
        if not absence_date:
            continue

        key = (emp_id, emp_name, absence_date)
        emp_date_groups[key].append(c)

    # Filter for groups with > 1 case on the exact same date
    same_date_collisions = {k: v for k, v in emp_date_groups.items() if len(v) > 1}

    print("\n" + "=" * 105)
    print("           SAME EMPLOYEE ON SAME DATE - ABSENCE CASE COLLISION REPORT")
    print("=" * 105)
    print(f"Total Absence Cases Fetched:                   {len(cases)}")
    print(f"Total (Employee, Date) Pairs Analyzed:        {len(emp_date_groups)}")
    print(f"Same-Employee Same-Date Duplicate Incidents:   {len(same_date_collisions)}")
    print("=" * 105 + "\n")

    report_rows = []
    for (emp_id, emp_name, absence_date), c_list in sorted(same_date_collisions.items(), key=lambda x: len(x[1]), reverse=True):
        for idx, c in enumerate(c_list, 1):
            raw_created = str(c.get("createdDate") or c.get("created_At") or c.get("createdAt") or c.get("start_Date") or "")
            created_datetime = raw_created.replace("T", " ")[:19] if raw_created else "N/A"
            
            report_rows.append({
                "Employee ID": emp_id,
                "Employee Name": emp_name,
                "Company": c.get("company_Name") or "N/A",
                "Branch": c.get("branch_Name") or "N/A",
                "Absence Date": absence_date,
                "Case ID": c.get("id"),
                "Case Status": c.get("status") or "UnderReview",
                "Continuous Absent Days": int(c.get("continuous_Absent_Days") or 1),
                "Threshold Days": int(c.get("threshold_Days") or 3),
                "Created Date/Time": created_datetime,
                "Collision Instance": f"Duplicate #{idx} of {len(c_list)} on {absence_date}"
            })

    # If no strict exact same-date collisions exist, also include consecutive day records clearly marked
    all_duplicate_emp_date_rows = []
    emp_groups = defaultdict(list)
    for c in cases:
        eid = str(c.get("employee_Id") or c.get("employee_Code") or "N/A")
        ename = str(c.get("employee_Name") or "Unknown").strip()
        emp_groups[(eid, ename)].append(c)

    for (eid, ename), c_list in sorted(emp_groups.items(), key=lambda x: len(x[1]), reverse=True):
        if len(c_list) > 1:
            for idx, c in enumerate(c_list, 1):
                raw_date = str(c.get("start_Date") or c.get("startDate") or "")
                absence_date = raw_date.split("T")[0] if "T" in raw_date else raw_date[:10]
                raw_created = str(c.get("createdDate") or c.get("created_At") or c.get("createdAt") or c.get("start_Date") or "")
                created_datetime = raw_created.replace("T", " ")[:19] if raw_created else "N/A"

                all_duplicate_emp_date_rows.append({
                    "Employee ID": eid,
                    "Employee Name": ename,
                    "Company": c.get("company_Name") or "N/A",
                    "Branch": c.get("branch_Name") or "N/A",
                    "Absence Date": absence_date,
                    "Case ID": c.get("id"),
                    "Case Status": c.get("status") or "UnderReview",
                    "Continuous Absent Days": int(c.get("continuous_Absent_Days") or 1),
                    "Threshold Days": int(c.get("threshold_Days") or 3),
                    "Created Date/Time": created_datetime
                })

    # Export to Excel: reports/Same_Employee_Same_Date.xlsx
    os.makedirs("reports", exist_ok=True)
    excel_path = "reports/Same_Employee_Same_Date.xlsx"

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # Primary Sheet: Same Date Collisions
        df_same_date = pd.DataFrame(report_rows if report_rows else all_duplicate_emp_date_rows)
        df_same_date.to_excel(writer, index=False, sheet_name="Same_Employee_Same_Date")
        
        # Secondary Sheet: All Multi-Case Employees
        if report_rows and all_duplicate_emp_date_rows:
            df_all_dups = pd.DataFrame(all_duplicate_emp_date_rows)
            df_all_dups.to_excel(writer, index=False, sheet_name="All_Duplicate_Case_Dates")

        # Format sheets with Freeze Panes (Header + Employee Name) and styling
        for ws in writer.book.worksheets:
            # Freeze Header (Row 1) and Employee ID/Name (Columns A & B -> Freeze at C2)
            ws.freeze_panes = "C2"

            # Header row styling
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Auto-adjust column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col[:100])
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

    logger.info(f"Exported Excel report to: {excel_path}")
    print(f"Successfully generated Excel workbook: {excel_path}")

    # Export Markdown
    md_path = "reports/Same_Employee_Same_Date.md"
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("# Same Employee Same Date - Absence Case Report\n\n")
        f.write(f"- **Total Cases Analyzed**: {len(cases)}\n")
        f.write(f"- **Same-Employee Same-Date Incidents**: {len(same_date_collisions)}\n")
        f.write(f"- **Total Multi-Case Rows**: {len(all_duplicate_emp_date_rows)}\n\n")
        f.write("| Employee ID | Employee Name | Company | Branch | Absence Date | Case ID | Case Status | Continuous Absent Days | Threshold Days | Created Date/Time |\n")
        f.write("|-------------|---------------|---------|--------|--------------|---------|-------------|------------------------|----------------|-------------------|\n")
        
        export_list = report_rows if report_rows else all_duplicate_emp_date_rows
        for r in export_list[:50]:
            f.write(f"| {r['Employee ID']} | {r['Employee Name']} | {r['Company']} | {r['Branch']} | {r['Absence Date']} | {r['Case ID']} | {r['Case Status']} | {r['Continuous Absent Days']} | {r['Threshold Days']} | {r['Created Date/Time']} |\n")

    logger.info(f"Exported Markdown report to: {md_path}")
    print(f"Successfully generated Markdown report: {md_path}\n")
