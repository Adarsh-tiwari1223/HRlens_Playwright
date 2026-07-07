"""
Payroll Validation: Excel vs API
================================
- Reads Excel using exact column indices (no header parsing ambiguity)
- Matches employees: primary = composite key (name_upper + code),
                     fallback = fuzzy name match (SequenceMatcher >= 0.85)
- Compares every mapped field column-by-column
- Outputs 1 multi-sheet Excel workbook:
    T0 Field Detail        — wide pivot: one row per employee, all fields as columns
    T1 Summary             — full counts: Excel, API, matched, fuzzy, unmatched
    T2 Difference Table    — salary match vs diff across ALL matched
    T3 Spelling Mismatch
    T4 Matched Exact
    T5 Difference
    T6 Not in Payroll      — in Excel, no API match
    T7 In API Not in Excel — in API, no Excel match
"""

import os
import re
import difflib
import pytest
import pandas as pd
from collections import defaultdict
from datetime import datetime
from utils.api.payroll_api import get_released_payroll

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
YEAR               = 2026
MONTH              = 4
PAYROLL_COMPANY_ID = 7
COMPANY            = "Originator Informatics Pvt. Ltd."
BRANCH             = "Varanasi (TEK)"
EXCEL_FILE         = "testdata/static/Originator-Salary-April'2026.xlsx"
SHEET_NAME         = "APR'26"
REPORTS_DIR        = "reports"
FUZZY_CUTOFF       = 0.85

# ---------------------------------------------------------------------------
# EXACT COLUMN INDICES  (0-indexed, rows 0-6 skipped)
# ---------------------------------------------------------------------------
COL_NAME = 4
COL_CODE = 7

# (Label, Excel_Col_Idx, Profile_Field, Release_Field)
VALIDATION_FIELDS = [
    ("Basic",          15,   "basic_salary",        "basic"),
    ("HRA",            16,   "hra",                 "hra"),
    ("Conveyance",     17,   "conveyance",          "conveyance_Allowance"),
    ("Incentive",      18,   None,                  "incentive"),
    ("Employee PF",    20,   "employeePF",          "employee_PF"),
    ("Employer PF",    None, "employerPF",          "employer_PF"),
    ("TDS",            23,   None,                  "tds"),
    ("Net Salary",     25,   "netTakeHomeSalary",   "netSalary"),
    ("Balance Leave",  26,   None,                  "balance_Leave"),
]

FIELD_HEADERS = {
    "Basic": ["Basic Excel", "Basic Profile", "Basic Payroll Release", "Basic Payroll Diff"],
    "HRA": ["HRA Excel", "HRA Profile", "HRA Payroll Release", "HRA Payroll Diff"],
    "Conveyance": ["Conveyance Excel", "Conveyance Profile", "Conveyance Payroll Release", "Conveyance Payroll Diff"],
    "Incentive": ["Incentive Excel", "Incentive Payroll Release", "Incentive Payroll Diff"],
    "Employee PF": ["Employee PF Excel", "Employee PF Profile", "Employee PF Payroll Release", "Employee PF Diff"],
    "Employer PF": ["Employer PF Excel", "Employer PF Profile", "Employer PF Payroll Release", "Employer PF Diff"],
    "TDS": ["TDS Excel", "TDS Payroll Release", "TDS Diff"],
    "Net Salary": ["Net Salary Excel", "Net Salary Profile", "Net Salary Payroll Release", "Net Salary Diff"],
    "Balance Leave": ["Balance Leave Excel", "Balance Leave Payroll Release", "Balance Leave Diff"],
}

# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def normalize_code(val) -> str:
    s = str(val).strip()
    return s[:-2] if s.endswith(".0") else s


def normalize_name(val) -> str:
    return str(val).strip().upper()


def safe_float(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, float) and pd.isna(val):
        return 0.0
    try:
        return round(float(str(val).strip()), 2)
    except (ValueError, TypeError):
        return 0.0


def is_blank(val) -> bool:
    if val is None:
        return True
    if isinstance(val, float) and pd.isna(val):
        return True
    return str(val).strip() == ""


# ---------------------------------------------------------------------------
# STEP 1: Read Excel
# ---------------------------------------------------------------------------

def read_excel(path: str, sheet: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet, header=None)


# ---------------------------------------------------------------------------
# STEP 2: Build API lookup
# ---------------------------------------------------------------------------

def build_api_lookup(api_records: list):
    composite_lookup, name_lookup = {}, {}
    for rec in api_records:
        name = normalize_name(rec.get("employeeName", ""))
        code = normalize_code(rec.get("employeeCode", ""))
        if name or code:
            composite_lookup[(name, code)] = rec
        if name:
            name_lookup[name] = rec
    return composite_lookup, name_lookup


# ---------------------------------------------------------------------------
# STEP 3: Match one Excel row to an API record
# ---------------------------------------------------------------------------

def find_api_record(excel_name, excel_code, composite_lookup, name_lookup):
    rec = composite_lookup.get((excel_name, excel_code))
    if rec:
        return rec, "EXACT_COMPOSITE"
    rec = name_lookup.get(excel_name)
    if rec:
        return rec, "EXACT_NAME"
    if excel_name:
        matches = difflib.get_close_matches(excel_name, list(name_lookup.keys()),
                                            n=1, cutoff=FUZZY_CUTOFF)
        if matches:
            return name_lookup[matches[0]], f"FUZZY({matches[0]})"
    return None, None


# Profile Cache to minimize API calls
_profile_cache = {}

def get_cached_profile(emp_id: int) -> dict:
    if emp_id not in _profile_cache:
        try:
            from utils.api.payroll_api import get_employee_detail
            _profile_cache[emp_id] = get_employee_detail(emp_id)
        except Exception:
            _profile_cache[emp_id] = {}
    return _profile_cache[emp_id]


# ---------------------------------------------------------------------------
# STEP 4: Compare fields for one matched employee row
# ---------------------------------------------------------------------------

def compare_fields(raw_row, excel_name, excel_code, api_rec, match_type, validation_fields_to_use=None) -> list:
    emp_id = api_rec.get("employeeId")
    profile_rec = get_cached_profile(emp_id) if emp_id else {}

    if validation_fields_to_use is None:
        validation_fields_to_use = VALIDATION_FIELDS

    results = []
    for label, excel_col, profile_field, release_field in validation_fields_to_use:
        # Excel Value
        if excel_col is not None:
            excel_raw = raw_row.iloc[excel_col] if excel_col < len(raw_row) else 0.0
            excel_val = safe_float(excel_raw)
        else:
            excel_val = 0.0

        # Profile Value
        if profile_field is not None:
            profile_val = safe_float(profile_rec.get(profile_field))
        else:
            profile_val = 0.0

        # Release Value
        release_val = safe_float(api_rec.get(release_field))

        # Diff Value (Excel - Payroll Release)
        diff_val = round(excel_val - release_val, 2)

        passed = (excel_val == release_val)

        results.append({
            "Employee Name": excel_name,
            "Employee Code": excel_code,
            "Match Type": match_type,
            "Field Name": label,
            "Excel Value": excel_val,
            "API Value": release_val,
            "Profile Value": profile_val,
            "Payroll Release Value": release_val,
            "Diff Value": diff_val,
            "Status": "Pass" if passed else "Fail",
        })
    return results


# ---------------------------------------------------------------------------
# STEP 5: Iterate all Excel rows
# ---------------------------------------------------------------------------

def compare_fields_no_excel(excel_name, excel_code, api_rec) -> list:
    emp_id = api_rec.get("employeeId")
    profile_rec = get_cached_profile(emp_id) if emp_id else {}

    results = []
    for label, excel_col, profile_field, release_field in VALIDATION_FIELDS:
        # Excel Value is NA
        excel_val = "NA"

        # Profile Value
        if profile_field is not None:
            profile_val = safe_float(profile_rec.get(profile_field))
        else:
            profile_val = 0.0

        # Release Value
        release_val = safe_float(api_rec.get(release_field))

        # Diff Value is NA
        diff_val = "NA"

        results.append({
            "Employee Name": excel_name,
            "Employee Code": excel_code,
            "Match Type": "—",
            "Field Name": label,
            "Excel Value": excel_val,
            "API Value": release_val,
            "Profile Value": profile_val,
            "Payroll Release Value": release_val,
            "Diff Value": diff_val,
            "Status": "Fail",
        })
    return results


def run_comparison(df, composite_lookup, name_lookup) -> list:
    detail_rows, matched_keys = [], set()

    if df is None or df.empty:
        # No Excel file, compare all API records with "NA" Excel values
        for (api_name, api_code), rec in composite_lookup.items():
            detail_rows.extend(compare_fields_no_excel(rec.get("employeeName", ""), api_code, rec))
        return detail_rows

    # 1. Detect Header Rows and Merge Consecutive Headers
    header_row_idx = -1
    for idx in range(min(15, len(df))):
        row = df.iloc[idx]
        row_vals = [str(val).strip().lower() for val in row if pd.notna(val)]
        has_name = any("name" in val for val in row_vals)
        has_code = any(any(x in val for x in ["code", "id", "emp.", "emp"]) for val in row_vals)
        if has_name and has_code:
            header_row_idx = idx
            break

    if header_row_idx == -1:
        header_row_idx = 5
        header_rows = [5]
    else:
        header_rows = [header_row_idx]
        if header_row_idx + 1 < min(15, len(df)):
            after_row = df.iloc[header_row_idx + 1]
            after_vals = [str(val).strip().lower() for val in after_row if pd.notna(val)]
            sub_terms = ["basic", "hra", "conv", "pf", "esi", "tds", "payable", "days", "leave", "deduction", "stipend"]
            if any(any(term in val for term in sub_terms) for val in after_vals):
                header_rows.append(header_row_idx + 1)

    num_cols = len(df.columns)
    merged_headers = []
    for c_idx in range(num_cols):
        parts = []
        for r_idx in header_rows:
            cell_val = df.iloc[r_idx, c_idx]
            if pd.notna(cell_val):
                cell_str = str(cell_val).replace("\n", " ").strip()
                cell_str = " ".join(cell_str.split())
                if cell_str and cell_str not in parts:
                    parts.append(cell_str)
        merged = " ".join(parts)
        merged_headers.append(" ".join(merged.split()).lower())

    # Detect Name and Code columns
    col_name = -1
    col_code = -1
    for idx, h in enumerate(merged_headers):
        if "name" in h and not any(x in h for x in ["father", "company", "bank", "branch"]):
            col_name = idx
            break
    for idx, h in enumerate(merged_headers):
        if any(x in h for x in ["code", "emp id", "employee id", "hrlenseid"]) and not "company" in h:
            col_code = idx
            break
    if col_code == -1:
        for idx, h in enumerate(merged_headers):
            if h == "emp" or h == "emp.":
                col_code = idx
                break

    if col_name == -1:
        col_name = COL_NAME
    if col_code == -1:
        col_code = COL_CODE

    # Map validation fields dynamically
    field_keywords = {
        "Basic": ["basic salary", "basic + da", "basic+da", "stipend", "basic"],
        "HRA": ["h.r.a. amt.", "hra amt.", "h.r.a.", "hra"],
        "Conveyance": ["conv. amt.", "conveyance", "conv.", "conveyance allowance"],
        "Incentive": ["incentive", "spcl. alw.", "special allowance", "special alw", "allowance", "incentives"],
        "Employee PF": ["deductions p.f.", "pf deduction", "employee pf", "pf", "p.f.", "esic & pf", "provident"],
        "Employer PF": ["employer pf", "employer p.f.", "employer_pf"],
        "TDS": ["deductions t.d.s.", "tds", "t.d.s."],
        "Net Salary": ["net payable", "net salary", "net amount", "total net", "net payable salary", "payable"],
        "Balance Leave": ["earned leave", "casual leave", "reserve leave", "balance leave", "cl", "el"]
    }

    local_validation_fields = []
    for label, excel_col_default, profile_field, release_field in VALIDATION_FIELDS:
        keywords = field_keywords.get(label, [])
        best_idx = -1
        for idx, h in enumerate(merged_headers):
            for kw in keywords:
                if kw == h or kw in h:
                    best_idx = idx
                    break
        if best_idx != -1:
            local_validation_fields.append((label, best_idx, profile_field, release_field))
        else:
            local_validation_fields.append((label, None, profile_field, release_field))

    start_row = header_rows[-1] + 1
    data_df = df.iloc[start_row:].reset_index(drop=True)

    for _, raw_row in data_df.iterrows():
        raw_name = raw_row.iloc[col_name] if col_name < len(raw_row) else None
        raw_code = raw_row.iloc[col_code] if col_code < len(raw_row) else None
        if is_blank(raw_name) and is_blank(raw_code):
            continue

        name_str = str(raw_name).strip().lower() if not is_blank(raw_name) else ""
        code_str = str(raw_code).strip().lower() if not is_blank(raw_code) else ""
        if any(x in name_str or x in code_str for x in ["total", "subtotal", "grand total", "dept total", "amount", "summary", "total:"]):
            continue

        excel_name = normalize_name(raw_name) if not is_blank(raw_name) else ""
        excel_code = normalize_code(raw_code) if not is_blank(raw_code) else ""
        api_rec, match_type = find_api_record(excel_name, excel_code,
                                              composite_lookup, name_lookup)

        if api_rec is None:
            detail_rows.append({
                "Employee Name": str(raw_name).strip(), "Employee Code": excel_code,
                "Match Type": "—", "Field Name": "—",
                "Excel Value": "—", "API Value": "—",
                "Status": "Not Found in API",
                "Remarks": "No API record matched (exact or fuzzy)",
            })
            continue

        api_key = (normalize_name(api_rec.get("employeeName", "")),
                   normalize_code(api_rec.get("employeeCode", "")))
        matched_keys.add(api_key)
        detail_rows.extend(compare_fields(raw_row, str(raw_name).strip(),
                                          excel_code, api_rec, match_type,
                                          validation_fields_to_use=local_validation_fields))

    for (api_name, api_code), rec in composite_lookup.items():
        if (api_name, api_code) not in matched_keys:
            detail_rows.append({
                "Employee Name": rec.get("employeeName", ""), "Employee Code": api_code,
                "Match Type": "—", "Field Name": "—",
                "Excel Value": "—", "API Value": "—",
                "Status": "Not Found in Excel",
                "Remarks": "Employee exists in API but not in Excel",
            })

    return detail_rows



# ---------------------------------------------------------------------------
# STEP 6A: Wide pivot — one row per employee, all fields as columns
# ---------------------------------------------------------------------------

def build_wide_rows(detail_rows: list):
    emp_fields, emp_match_type, emp_order = defaultdict(dict), {}, []

    for r in detail_rows:
        if r["Status"] in ("Not Found in API", "Not Found in Excel"):
            continue
        key = (r["Employee Name"], r["Employee Code"])
        if key not in emp_fields:
            emp_order.append(key)
        emp_fields[key][r["Field Name"]] = r
        emp_match_type[key] = r["Match Type"]

    field_labels = [f[0] for f in VALIDATION_FIELDS]

    wide_rows = []
    for key in emp_order:
        name, code = key
        row = {
            "Employee Name": name,
            "Employee Code": code,
            "Match Type":    emp_match_type.get(key, "—"),
        }
        any_fail = False
        for label in field_labels:
            fr = emp_fields[key].get(label)
            headers = FIELD_HEADERS[label]
            if fr:
                excel_v = fr["Excel Value"]
                profile_v = fr["Profile Value"]
                release_v = fr["Payroll Release Value"]
                diff_v = fr["Diff Value"]
                if fr["Status"] == "Fail":
                    any_fail = True
            else:
                excel_v = profile_v = release_v = diff_v = "—"

            if len(headers) == 4:
                excel_col_name, profile_col_name, release_col_name, diff_col_name = headers
                row[excel_col_name] = excel_v
                row[profile_col_name] = profile_v
                row[release_col_name] = release_v
                row[diff_col_name] = diff_v
            else:
                excel_col_name, release_col_name, diff_col_name = headers
                row[excel_col_name] = excel_v
                row[release_col_name] = release_v
                row[diff_col_name] = diff_v

        row["Final Status"] = "Fail" if any_fail else "Pass"
        wide_rows.append(row)

    return wide_rows, field_labels


# ---------------------------------------------------------------------------
# STEP 6B: Build report tables from detail_rows
# ---------------------------------------------------------------------------

def build_tables(detail_rows: list, api_total: int = 0):
    exact_paid, exact_diff = [], []
    fuzzy_paid, fuzzy_diff = [], []
    spelling_rows, not_found, not_in_excel = [], [], []
    seen_spelling = set()
    emp_fields, emp_match_type = defaultdict(dict), {}

    for r in detail_rows:
        if r["Status"] == "Not Found in API":
            not_found.append({
                "Employee Name": r["Employee Name"],
                "Employee Code": r["Employee Code"],
                "Parent Salary": "—",
            })
            continue
        if r["Status"] == "Not Found in Excel":
            not_in_excel.append({
                "Employee Name": r["Employee Name"],
                "Employee Code": r["Employee Code"],
            })
            continue
        key = (r["Employee Name"], r["Employee Code"])
        emp_fields[key][r["Field Name"]] = r
        emp_match_type[key] = r["Match Type"]

    for key, fields in emp_fields.items():
        name, code = key
        match_type = emp_match_type[key]
        is_fuzzy   = "FUZZY" in match_type
        net_row    = fields.get("Net Salary")
        excel_net  = net_row["Excel Value"] if net_row else "—"
        api_net    = net_row["API Value"]   if net_row else "—"
        net_match  = net_row and net_row["Status"] == "Pass"

        entry = {"Employee Name": name, "Employee Code": code,
                 "Parent Salary": excel_net, "Payroll Salary": api_net}

        if is_fuzzy:
            if key not in seen_spelling:
                seen_spelling.add(key)
                m = re.search(r"FUZZY\((.+)\)", match_type)
                spelling_rows.append({
                    "Parent Column (Excel)": name,
                    "Payroll Column (API)":  m.group(1) if m else match_type,
                    "Type": "Spelling",
                })
            if net_match:
                fuzzy_paid.append(entry)
            else:
                try:
                    diff = round(float(excel_net) - float(api_net), 2)
                except (TypeError, ValueError):
                    diff = "—"
                fuzzy_diff.append({**entry, "Difference": diff})
        else:
            if net_match:
                exact_paid.append(entry)
            else:
                try:
                    diff = round(float(excel_net) - float(api_net), 2)
                except (TypeError, ValueError):
                    diff = "—"
                exact_diff.append({**entry, "Difference": diff})

    n_exact          = len(exact_paid) + len(exact_diff)
    n_fuzzy          = len(seen_spelling)
    n_total_matched  = len(emp_fields)
    n_salary_match  = len(exact_paid)
    n_salary_diff   = len(exact_diff)
    n_not_api       = len(not_found)
    n_not_excel     = len(not_in_excel)
    n_excel_total   = n_total_matched + n_not_api
    n_api_total     = api_total

    t1 = [
        {"Category": "Excel Records (Source)",     "Count": n_excel_total},
        {"Category": "API Records (Payroll)",       "Count": n_api_total},
        {"Category": "---",                         "Count": "---"},
        {"Category": "Matched (Exact)",             "Count": n_exact},
        {"Category": "Total Matched",               "Count": n_exact},
        {"Category": "Matched (Fuzzy / Spelling)",  "Count": n_fuzzy},
        {"Category": "---",                         "Count": "---"},
        {"Category": "Salary Match (= Paid)",       "Count": n_salary_match},
        {"Category": "Salary Diff  (!= Paid)",      "Count": n_salary_diff},
        {"Category": "---",                         "Count": "---"},
        {"Category": "Unmatched (Not in API)",      "Count": n_not_api},
        {"Category": "Unmatched (Not in Excel)",    "Count": n_not_excel},
        {"Category": "Total Unmatched",             "Count": n_not_api + n_not_excel},
    ]

    t2 = [
        {"Status": "Salary Match (Exact)",  "Count": len(exact_paid)},
        {"Status": "Salary Diff  (Exact)",  "Count": len(exact_diff)},
        {"Status": "---",                   "Count": "---"},
        {"Status": "Total Salary Match",    "Count": n_salary_match},
        {"Status": "Total Salary Diff",     "Count": n_salary_diff},
        {"Status": "Total Matched",         "Count": n_exact},
    ]

    all_diff = exact_diff + fuzzy_diff

    for i, r in enumerate(spelling_rows, 1): r["#"] = i
    for i, r in enumerate(exact_paid,    1): r["#"] = i
    for i, r in enumerate(all_diff,      1): r["#"] = i
    for i, r in enumerate(not_found,     1): r["#"] = i
    for i, r in enumerate(not_in_excel,  1): r["#"] = i

    return t1, t2, spelling_rows, exact_paid, all_diff, not_found, not_in_excel


# ---------------------------------------------------------------------------
# STEP 7: Save single multi-sheet Excel workbook
# ---------------------------------------------------------------------------

def save_workbook(wide_rows, field_labels, t1, t2, t3, t4, t5, t6, t7, custom_path=None, company_name=None) -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(REPORTS_DIR, exist_ok=True)
    if custom_path:
        path = custom_path
    else:
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = f"{REPORTS_DIR}/payroll_validation_report_{ts}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    MONTH_NAME   = datetime(YEAR, MONTH, 1).strftime("%B %Y").upper()
    c_name = company_name if company_name else COMPANY
    REPORT_TITLE = f"PAYROLL MAPPING REPORT  —  {c_name}  —  {BRANCH}  —  {MONTH_NAME}"

    HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(bold=True, size=13)
    SEP_FILL   = PatternFill("solid", fgColor="D9E1F2")
    FAIL_FILL  = PatternFill("solid", fgColor="FFE0E0")
    PASS_FILL  = PatternFill("solid", fgColor="E2EFDA")
    WARN_FILL  = PatternFill("solid", fgColor="FFF2CC")
    thin       = Side(style="thin")
    bdr        = Border(left=thin, right=thin, top=thin, bottom=thin)

    def make_sheet(title):
        ws = wb.create_sheet(title=title)
        ws.append([REPORT_TITLE])
        ws["A1"].font = TITLE_FONT
        ws.append([])
        return ws

    def write_header(ws, cols):
        ws.append(cols)
        row = ws.max_row
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = bdr

    def write_row(ws, values, fill=None):
        ws.append(values)
        row = ws.max_row
        for c in range(1, len(values) + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = bdr
            cell.alignment = Alignment(horizontal="left")
            if fill:
                cell.fill = fill

    def write_total(ws, count):
        ws.append([])
        ws.append(["Total:", count])
        row = ws.max_row
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)

    def autofit(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 50)

    # ── T0: Wide Field Detail ─────────────────────────────────────────────
    ws0 = make_sheet("T0 Field Detail")
    ws0.append(["TABLE 0: FIELD-LEVEL DETAIL (Wide View)"])
    ws0[f"A{ws0.max_row}"].font = Font(bold=True, size=12)

    fixed_cols = ["Employee Name", "Employee Code", "Match Type"]
    field_cols = []
    for label in field_labels:
        field_cols += FIELD_HEADERS[label]
    all_cols = fixed_cols + field_cols + ["Final Status"]
    write_header(ws0, all_cols)

    final_status_col = len(all_cols)
    data_start_row   = 5

    for r in wide_rows:
        values = [r["Employee Name"], r["Employee Code"], r["Match Type"]]
        for label in field_labels:
            headers = FIELD_HEADERS[label]
            if len(headers) == 4:
                excel_col_name, profile_col_name, release_col_name, diff_col_name = headers
                values += [r[excel_col_name], r[profile_col_name], r[release_col_name], r[diff_col_name]]
            else:
                excel_col_name, release_col_name, diff_col_name = headers
                values += [r[excel_col_name], r[release_col_name], r[diff_col_name]]
        values.append(r["Final Status"])
        write_row(ws0, values)

    # Find indices of all diff columns (1-indexed for openpyxl)
    diff_col_indices = []
    for col_idx, col_name in enumerate(all_cols, 1):
        if col_name.endswith("Diff"):
            diff_col_indices.append(col_idx)

    for row_idx in range(data_start_row, ws0.max_row + 1):
        # Color final status
        status_cell = ws0.cell(row=row_idx, column=final_status_col)
        if status_cell.value == "Pass":
            status_cell.fill = PASS_FILL
        elif status_cell.value == "Fail":
            status_cell.fill = FAIL_FILL

        # Color non-zero diff columns
        for diff_col_idx in diff_col_indices:
            diff_cell = ws0.cell(row=row_idx, column=diff_col_idx)
            try:
                val = float(diff_cell.value)
                if abs(val) > 0.001:
                    diff_cell.fill = FAIL_FILL
                    diff_cell.font = Font(bold=True, color="9C0006")
            except (ValueError, TypeError):
                pass

    autofit(ws0)

    # ── T1: Summary ───────────────────────────────────────────────────────
    ws1 = make_sheet("T1 Summary")
    ws1.append(["TABLE 1: SUMMARY"])
    ws1[f"A{ws1.max_row}"].font = Font(bold=True, size=12)
    write_header(ws1, ["Category", "Count"])
    for r in t1:
        is_sep  = r["Category"] == "---"
        is_bold = r["Category"].startswith("Total") or r["Category"].startswith("Excel") or r["Category"].startswith("API")
        write_row(ws1, [r["Category"], r["Count"]], SEP_FILL if is_sep else None)
        if is_bold:
            row = ws1.max_row
            ws1.cell(row=row, column=1).font = Font(bold=True)
            ws1.cell(row=row, column=2).font = Font(bold=True)
    autofit(ws1)

    # ── T2: Difference Table ──────────────────────────────────────────────
    ws2 = make_sheet("T2 Difference Table")
    ws2.append(["TABLE 2: DIFFERENCE TABLE"])
    ws2[f"A{ws2.max_row}"].font = Font(bold=True, size=12)
    write_header(ws2, ["Status", "Count"])
    for r in t2:
        is_sep  = r["Status"] == "---"
        is_bold = r["Status"].startswith("Total")
        write_row(ws2, [r["Status"], r["Count"]], SEP_FILL if is_sep else None)
        if is_bold:
            row = ws2.max_row
            ws2.cell(row=row, column=1).font = Font(bold=True)
            ws2.cell(row=row, column=2).font = Font(bold=True)
    autofit(ws2)

    # ── T3: Spelling Mismatch ─────────────────────────────────────────────
    ws3 = make_sheet("T3 Spelling Mismatch")
    ws3.append(["TABLE 3: SPELLING MISMATCH (Fuzzy Matched)"])
    ws3[f"A{ws3.max_row}"].font = Font(bold=True, size=12)
    write_header(ws3, ["#", "Parent Column (Excel)", "Payroll Column (API)", "Type"])
    for r in t3:
        write_row(ws3, [r["#"], r["Parent Column (Excel)"], r["Payroll Column (API)"], r["Type"]])
    write_total(ws3, len(t3))
    autofit(ws3)

    # ── T4: Matched Exact (salary match, non-fuzzy) ───────────────────────
    ws4 = make_sheet("T4 Matched Exact")
    ws4.append(["TABLE 4: MATCHED (Exact — Salary Match)"])
    ws4[f"A{ws4.max_row}"].font = Font(bold=True, size=12)
    write_header(ws4, ["#", "Employee Name", "Employee Code", "Parent Salary", "Payroll Salary"])
    for r in t4:
        write_row(ws4, [r["#"], r["Employee Name"], r["Employee Code"],
                        r["Parent Salary"], r["Payroll Salary"]])
    write_total(ws4, len(t4))
    autofit(ws4)

    # ── T5: All Salary Differences (Wide View for Fails) ──────────────────
    ws5 = make_sheet("T5 Difference")
    ws5.append(["TABLE 5: DIFFERENCE (Failed Matches)"])
    ws5[f"A{ws5.max_row}"].font = Font(bold=True, size=12)

    write_header(ws5, all_cols)

    fail_rows = [r for r in wide_rows if r["Final Status"] == "Fail"]
    for r in fail_rows:
        values = [r["Employee Name"], r["Employee Code"], r["Match Type"]]
        for label in field_labels:
            headers = FIELD_HEADERS[label]
            if len(headers) == 4:
                excel_col_name, profile_col_name, release_col_name, diff_col_name = headers
                values += [r[excel_col_name], r[profile_col_name], r[release_col_name], r[diff_col_name]]
            else:
                excel_col_name, release_col_name, diff_col_name = headers
                values += [r[excel_col_name], r[release_col_name], r[diff_col_name]]
        values.append(r["Final Status"])
        write_row(ws5, values)

    for row_idx in range(data_start_row, ws5.max_row + 1):
        # Color final status
        status_cell = ws5.cell(row=row_idx, column=final_status_col)
        if status_cell.value == "Pass":
            status_cell.fill = PASS_FILL
        elif status_cell.value == "Fail":
            status_cell.fill = FAIL_FILL

        # Color non-zero diff columns
        for diff_col_idx in diff_col_indices:
            diff_cell = ws5.cell(row=row_idx, column=diff_col_idx)
            try:
                val = float(diff_cell.value)
                if abs(val) > 0.001:
                    diff_cell.fill = FAIL_FILL
                    diff_cell.font = Font(bold=True, color="9C0006")
            except (ValueError, TypeError):
                pass

    autofit(ws5)

    # ── T6: Not in Payroll (Excel → no API match) ─────────────────────────
    ws6 = make_sheet("T6 Not in Payroll")
    ws6.append(["TABLE 6: NOT IN PAYROLL (In Excel, No API Match)"])
    ws6[f"A{ws6.max_row}"].font = Font(bold=True, size=12)
    write_header(ws6, ["#", "Employee Name", "Employee Code", "Parent Salary"])
    for r in t6:
        write_row(ws6, [r["#"], r["Employee Name"], r["Employee Code"], r["Parent Salary"]])
    write_total(ws6, len(t6))
    autofit(ws6)

    # ── T7: In API but Not in Excel ───────────────────────────────────────
    ws7 = make_sheet("T7 In API Not in Excel")
    ws7.append(["TABLE 7: IN API BUT NOT IN EXCEL"])
    ws7[f"A{ws7.max_row}"].font = Font(bold=True, size=12)
    write_header(ws7, ["#", "Employee Name (API)", "Employee Code (API)"])
    for r in t7:
        write_row(ws7, [r["#"], r["Employee Name"], r["Employee Code"]], WARN_FILL)
    write_total(ws7, len(t7))
    autofit(ws7)

    wb.save(path)
    print(f"\nReport saved: {path}")
    return path


# ---------------------------------------------------------------------------
# PYTEST ENTRY POINT
# ---------------------------------------------------------------------------

@pytest.mark.api
@pytest.mark.regression
def test_payroll_excel_vs_api_field_level():
    """
    Field-level payroll validation — Excel vs API.
    Matching: exact composite key first, fuzzy name fallback.
    Asserts zero Net Payable mismatches; saves 1 multi-sheet Excel report.
    """
    try:
        api_response = get_released_payroll(YEAR, MONTH, PAYROLL_COMPANY_ID)
    except Exception as exc:
        pytest.fail(f"API call failed: {exc}")

    api_records = api_response.get("data", [])
    assert api_records, "API returned an empty data list"

    if not os.path.exists(EXCEL_FILE):
        pytest.skip(f"Excel file not found: {EXCEL_FILE}")

    df = read_excel(EXCEL_FILE, SHEET_NAME)
    composite_lookup, name_lookup = build_api_lookup(api_records)
    detail_rows = run_comparison(df, composite_lookup, name_lookup)

    wide_rows, field_labels = build_wide_rows(detail_rows)
    t1, t2, t3, t4, t5, t6, t7 = build_tables(detail_rows, api_total=len(api_records))
    report_path = save_workbook(wide_rows, field_labels, t1, t2, t3, t4, t5, t6, t7)

    print(f"\n{'='*58}")
    print(f"  PAYROLL VALIDATION — {YEAR}-{MONTH:02d}")
    print(f"{'='*58}")
    print(f"  Excel Records        : {t1[0]['Count']}")
    print(f"  API Records          : {t1[1]['Count']}")
    print(f"  Total Matched        : {t1[4]['Count']}")
    print(f"    Matched (Exact)    : {t1[3]['Count']}")
    print(f"    Matched (Fuzzy)    : {t1[5]['Count']}")
    print(f"  Salary Match         : {t1[7]['Count']}")
    print(f"  Salary Diff          : {t1[8]['Count']}")
    print(f"  Not in API           : {t1[10]['Count']}")
    print(f"  Not in Excel         : {t1[11]['Count']}")
    print(f"  Total Unmatched      : {t1[12]['Count']}")
    print(f"{'='*58}")
    print(f"  Report: {report_path}")

    fail_rows = [r for r in wide_rows if r["Final Status"] == "Fail"]
    if fail_rows:
        print(f"\nDifferences (first 30 of {len(fail_rows)}):")
        for r in fail_rows[:30]:
            mismatches = []
            for label in field_labels:
                headers = FIELD_HEADERS[label]
                if len(headers) == 4:
                    excel_col_name, _, release_col_name, _ = headers
                else:
                    excel_col_name, release_col_name, _ = headers
                excel_val = r[excel_col_name]
                release_val = r[release_col_name]
                if excel_val != release_val:
                    mismatches.append(f"{label}: Excel={excel_val} API={release_val}")
            print(f"  [{r['Employee Code']}] {r['Employee Name']:<30} | {', '.join(mismatches)}")

    assert not fail_rows, (
        f"{len(fail_rows)} employee(s) have field mismatches. "
        "See T5 Difference sheet for details."
    )
