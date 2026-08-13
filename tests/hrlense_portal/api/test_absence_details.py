"""
Pure API Test Suite: Absence Details Retrieval (Read-Only GET).
Does NOT launch a browser — executes direct API requests.
"""

import os
import json
import logging
import pytest
import pandas as pd
from core.config import settings
from utils.api.absence_api import get_absence_case_list

logger = logging.getLogger(__name__)


@pytest.mark.api
@pytest.mark.attendance
def test_get_absence_details_api_only():
    """
    STRICTLY READ-ONLY (GET ONLY via API):
    Fetches absence case records directly from:
    GET /api/AbsenceCase/list
    Formats employee details and exports directly to Excel and Markdown.
    """
    logger.info(f"Fetching Absence Details (API ONLY) from {settings.API_BASE_URL}...")

    # Direct API Call (No Browser)
    records = get_absence_case_list(rows=1000, search="", status="")

    print("\n" + "=" * 135)
    print(f"               HR LENS - ABSENCE DETAILS REPORT (API ONLY) | ENV: {settings.ENV.upper()}")
    print("=" * 135)
    print(f"Total Absence Cases Fetched: {len(records)}\n")

    assert len(records) > 0, "Expected at least 1 absence record from API."

    parsed_rows = []
    for item in records:
        case_id = item.get("id", "")
        emp_name = item.get("employee_Name") or item.get("employeeName") or item.get("name") or "Unknown"
        emp_code = item.get("employee_Code") or item.get("employeeCode") or item.get("employee_Id") or item.get("employeeId") or "N/A"
        emp_id = item.get("employee_Id") or item.get("employeeId") or ""
        branch = item.get("branch_Name") or item.get("branchName") or "N/A"
        company = item.get("company_Name") or item.get("companyName") or "N/A"
        dept = item.get("department_Name") or item.get("departmentName") or "N/A"
        raw_date = item.get("start_Date") or item.get("startDate") or ""
        start_date = raw_date.split("T")[0] if "T" in str(raw_date) else str(raw_date)[:10]
        absent_days = item.get("continuous_Absent_Days") or item.get("continuousAbsentDays") or 1
        threshold = item.get("threshold_Days") or item.get("thresholdDays") or 3
        manager = item.get("manager_Name") or item.get("managerName") or "N/A"
        tl = item.get("tl_Name") or item.get("tlName") or "N/A"
        status = item.get("status") or item.get("caseStatus") or "UnderReview"
        comm_count = item.get("communication_Count") or 0
        raw_comm_date = item.get("last_Communication_Date") or ""
        last_comm_date = raw_comm_date.split("T")[0] if "T" in str(raw_comm_date) else str(raw_comm_date)[:10]

        parsed_rows.append({
            "Case ID": case_id,
            "Employee Name": emp_name,
            "Emp Code": emp_code,
            "Emp ID": emp_id,
            "Branch": branch,
            "Company": company,
            "Department": dept,
            "Start Date": start_date,
            "Continuous Absent Days": absent_days,
            "Threshold Days": threshold,
            "Manager": manager,
            "Team Lead": tl,
            "Status": status,
            "Communication Count": comm_count,
            "Last Communication Date": last_comm_date
        })

    # Print Table Preview
    print(f"{'Emp Name':<22} | {'Code':<6} | {'Branch':<12} | {'Company':<12} | {'Start Date':<10} | {'Absent Days':<11} | {'Manager':<18} | {'Status':<12}")
    print("-" * 135)

    for r in parsed_rows[:35]:
        print(f"{r['Employee Name'][:22]:<22} | {str(r['Emp Code'])[:6]:<6} | {r['Branch'][:12]:<12} | {r['Company'][:12]:<12} | {r['Start Date']:<10} | {str(r['Continuous Absent Days']):<11} | {r['Manager'][:18]:<18} | {r['Status']:<12}")

    print("-" * 135)
    print(f"Displaying top 35 of {len(parsed_rows)} records in console. Full dataset exported to Excel.")
    print("=" * 135 + "\n")

    # Export to Excel
    os.makedirs("reports", exist_ok=True)
    excel_path = "reports/absence_details.xlsx"
    df = pd.DataFrame(parsed_rows)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Absence_Details")
        ws = writer.sheets["Absence_Details"]
        
        # Freeze Top Row and Employee Name column (Column B -> Freeze at C2)
        ws.freeze_panes = "C2"

        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    logger.info(f"Exported Absence Details Excel report with Freeze Panes to: {excel_path}")
    print(f"Successfully generated Excel report: {excel_path}")

    # Export to Markdown
    md_path = "reports/absence_details.md"
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(f"# Absence Details Report (API ONLY - {settings.ENV.upper()})\n\n")
        f.write("| Employee Name | Emp Code | Branch | Company | Department | Start Date | Absent Days | Manager | Status |\n")
        f.write("|---------------|----------|--------|---------|------------|------------|-------------|---------|--------|\n")
        for r in parsed_rows:
            f.write(f"| {r['Employee Name']} | {r['Emp Code']} | {r['Branch']} | {r['Company']} | {r['Department']} | {r['Start Date']} | {r['Continuous Absent Days']} | {r['Manager']} | {r['Status']} |\n")

    logger.info(f"Exported Absence Details Markdown report to: {md_path}")
