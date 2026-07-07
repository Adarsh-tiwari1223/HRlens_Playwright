import os
import glob
import openpyxl

def inspect():
    path = "reports/payroll blocker report/*.xlsx"
    files = glob.glob(path)
    print(f"Found {len(files)} blocker Excel files:")
    for f in files:
        print(f"\nFile: {os.path.basename(f)}")
        try:
            wb = openpyxl.load_workbook(f, read_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = list(ws.iter_rows(values_only=True))
                print(f"  Sheet '{sheet}': {len(rows)} rows")
                if len(rows) > 0:
                    print(f"    Row 1 (Title): {rows[0][:2]}")
                if len(rows) > 2:
                    print(f"    Row 3 (Headers): {rows[2][:5]}")
                if len(rows) > 3:
                    print(f"    Row 4 (First Data): {rows[3][:5]}")
                    print(f"    ... and {len(rows)-4} more rows")
            wb.close()
        except Exception as e:
            print(f"  Error reading: {e}")

if __name__ == "__main__":
    inspect()
